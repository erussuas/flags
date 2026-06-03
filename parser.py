"""
parser.py — EnergyCAP Report-27 Bill Flags parser
Converts the xlsx text dump into a clean pandas DataFrame.
"""
import re
import shutil
import subprocess
import tempfile
import os
from collections import Counter
from datetime import datetime

import pandas as pd


# ── Flag metadata ─────────────────────────────────────────────────────────────
FLAG_META = {
    "Rate schedule mismatch": {
        "category": "Import / Configuration", "priority": "Medium",
        "cause": "Rate schedule in import file doesn't match the rate schedule assigned to the meter in EnergyCAP.",
        "action": "Update the meter's rate schedule in EnergyCAP to match the utility's current rate, or correct the import file if the utility rate is wrong.",
        "in_energycap": "Navigate to the meter record → Rate Schedule field and update to the imported value if correct.",
    },
    "Serial number mismatch": {
        "category": "Import / Configuration", "priority": "Medium",
        "cause": "Serial number in the import file doesn't match the serial number stored in EnergyCAP. May indicate the vendor swapped the physical meter.",
        "action": "Verify with the vendor whether the meter was replaced. If replaced, update the serial number in EnergyCAP. If it's a data entry error, correct the import file.",
        "in_energycap": "Navigate to meter record → Serial Number field. If meter was swapped, retire old meter and create a new one.",
    },
    "Duplicate bill": {
        "category": "Duplicate / Overlap", "priority": "High",
        "cause": "Total bill cost equals a prior bill's cost AND start/end dates match — likely a data entry or billing error.",
        "action": "Compare with the prior bill. If truly duplicate, void one. If it's a legitimate re-bill, confirm with the vendor and document the reason.",
        "in_energycap": "Open both bills, compare line items. Void the duplicate via More Actions → Void Bill.",
    },
    "Overlapping bill": {
        "category": "Duplicate / Overlap", "priority": "High",
        "cause": "One or more bills have overlapping start/end dates on the same account.",
        "action": "Review dates of all overlapping bills. Correct dates if it's a data entry error, or contact vendor if dates were issued incorrectly.",
        "in_energycap": "Use the Account History tab on the bill to view adjacent bills and identify the overlap.",
    },
    "Multiple bills in period": {
        "category": "Duplicate / Overlap", "priority": "High",
        "cause": "More than one bill exists for the same account and billing period.",
        "action": "Determine if these are split bills (valid) or accidental duplicates. Void any true duplicates. If split billing is expected, consider suppressing this audit for that account.",
        "in_energycap": "Filter bill list by account and period. Void duplicates via More Actions → Void Bill.",
    },
    "Abnormal cost": {
        "category": "Statistical Outlier", "priority": "High",
        "cause": "Total cost is a severe statistical outlier (>3.0 standard deviations) vs last 12 bills using quadratic regression.",
        "action": "Review bill line items for unexpected charges. Check if a rate change occurred. Contact vendor if cost increase is unexplained. Consider cost recovery if an error is found.",
        "in_energycap": "Open bill → check line items for unusual charges. Use Account History tab to compare to prior bills.",
    },
    "Abnormal use": {
        "category": "Statistical Outlier", "priority": "High",
        "cause": "Usage is a severe statistical outlier compared to the last 12 bills. Could indicate meter issue, leak, or billing error.",
        "action": "Check for operational changes at the site (new equipment, process change, leak). Contact vendor if usage spike is unexplained. If meter issue suspected, request re-read.",
        "in_energycap": "Open bill → review meter use values. Compare to Account History. Send notification to site manager.",
    },
    "Abnormal demand": {
        "category": "Statistical Outlier", "priority": "High",
        "cause": "Demand reading is a severe statistical outlier compared to historical bills.",
        "action": "Investigate if new high-load equipment was added. Verify demand meter readings with the vendor. Check if ratchet clauses may create ongoing cost impact.",
        "in_energycap": "Review demand meter readings and compare to prior months via Account History.",
    },
    "Gap between bills": {
        "category": "Date / Timeline", "priority": "Medium",
        "cause": "Gap of 2 or more days between this bill and the preceding bill — could mean a missing bill.",
        "action": "Check if a bill was skipped or not yet received. Contact vendor or check vendor portal for missing invoices. Use Report-17 to find 1-day gaps.",
        "in_energycap": "Use Account History tab to see the gap. If a bill is missing, manually enter it or re-import.",
    },
    "Shorter or longer bill": {
        "category": "Date / Timeline", "priority": "Low",
        "cause": "Bill period is significantly shorter or longer than the average of prior bills. Could be a meter read cycle change or data entry error.",
        "action": "Verify with the vendor that the billing period is correct. If dates are off, correct them. If cycle changed, this flag may recur — consider suppressing for the account.",
        "in_energycap": "Review Start and End Date on the bill header. Correct if needed.",
    },
    "Late statement date": {
        "category": "Date / Timeline", "priority": "Low",
        "cause": "Statement date is too many days after the bill's end date — possible data entry error or delayed invoice.",
        "action": "Verify the statement date is correct. If the vendor consistently sends invoices late, consider adjusting the audit threshold for this vendor.",
        "in_energycap": "Correct the Statement Date in the bill header if it was entered incorrectly.",
    },
    "Late due date": {
        "category": "Date / Timeline", "priority": "Low",
        "cause": "Due date is too many days after the bill's end date.",
        "action": "Verify due date accuracy to avoid late payment penalties. Correct if it was a data entry error.",
        "in_energycap": "Correct the Due Date in the bill header.",
    },
    "Unexpected billing period": {
        "category": "Date / Timeline", "priority": "Medium",
        "cause": "Billing period is outside the bill's start and end dates — likely a data entry error.",
        "action": "Correct the billing period to align with the start and end dates of the bill.",
        "in_energycap": "Edit the bill header and set the Billing Period to match the start/end dates.",
    },
    "Flagged line item type found": {
        "category": "Line Item Review", "priority": "Medium",
        "cause": "Bill contains a line item type configured for monitoring (e.g., late fees, taxes, penalties).",
        "action": "Review the specific line item. If it's a late fee, confirm the prior bill was paid on time. If it's an unexpected charge, dispute with vendor.",
        "in_energycap": "Open bill → scroll to line items section → review the flagged line item type.",
    },
    "Flagged line item description found": {
        "category": "Line Item Review", "priority": "Medium",
        "cause": "Bill contains a line item with a description matching a monitored keyword (case-insensitive partial match).",
        "action": "Review the line item description and value. Common triggers: 'balance forward', 'prior balance', 'penalty'. Confirm whether charge is legitimate.",
        "in_energycap": "Open bill → check line items for the matching description.",
    },
    "High use per day": {
        "category": "Statistical Outlier", "priority": "Medium",
        "cause": "Use per day exceeds 3× the highest use-per-day from the top 90% of prior bills (last 4 years).",
        "action": "Investigate operational changes or equipment issues at the site. Verify with vendor if meter read is accurate.",
        "in_energycap": "Compare current use/day to historical via Account History tab.",
    },
    "High cost per day": {
        "category": "Statistical Outlier", "priority": "Medium",
        "cause": "Cost per day exceeds 3× the highest cost-per-day from the top 90% of prior bills.",
        "action": "Check for rate changes, new charges, or usage spikes driving up cost. Compare bill line items to prior periods.",
        "in_energycap": "Open bill and compare line-by-line with a prior period bill.",
    },
    "Conflicting use units": {
        "category": "Import / Configuration", "priority": "Medium",
        "cause": "The use unit on the bill (e.g. CCF) conflicts with the use unit on the meter in EnergyCAP (e.g. Therms).",
        "action": "Determine the correct unit of measure. Update the meter configuration in EnergyCAP if it's wrong, or correct the import file if the vendor changed reporting units.",
        "in_energycap": "Navigate to meter record → Use Unit field. Update to match the vendor's unit if the vendor changed it.",
    },
}

PRIORITY_ORDER = {"High": 0, "Medium": 1, "Low": 2}
PRIORITY_COLOR = {"High": "#dc3545", "Medium": "#fd7e14", "Low": "#198754"}
CATEGORY_COLOR = {
    "Import / Configuration": "#0d6efd",
    "Duplicate / Overlap": "#dc3545",
    "Statistical Outlier": "#6f42c1",
    "Date / Timeline": "#fd7e14",
    "Line Item Review": "#198754",
}


# ── Parser ────────────────────────────────────────────────────────────────────

def parse_report27_text(text: str) -> pd.DataFrame:
    """
    Parse the tab-separated text dump of a Report-27 xlsx into a DataFrame.
    Each row = one flagged bill.
    """
    lines = text.split("\n")
    records: list[dict] = []
    current: dict = {}

    for line in lines:
        stripped = line.strip()

        # ── New account block ─────────────────────────────────────────────
        if re.search(r'Account:\s{2,}', stripped):
            if current.get("bill_id"):
                records.append(current)
            current = {
                "account": re.sub(r'Account:\s+', '', stripped).strip(),
                "address": "", "vendor": "", "bill_id": "",
                "billing_period": "", "cost": 0.0,
                "status": "", "flag_issues": "", "assigned_to": "",
                "cost_recovery": 0.0,
                "flagged_date": None, "resolved_date": None,
                "num_issues": 0, "resolvers": [], "flag_events": [],
            }
            continue

        # ── Address ───────────────────────────────────────────────────────
        if (current.get("account") and not current.get("address")
                and not current.get("vendor") and stripped
                and "Vendor:" not in stripped
                and not re.match(r'\d{6}', stripped)):
            current["address"] = stripped
            continue

        # ── Vendor ────────────────────────────────────────────────────────
        if "Vendor:" in stripped:
            vm = re.search(r'Vendor:\s+(.+?)(?:\s*\[|$)', stripped)
            if vm:
                current["vendor"] = vm.group(1).strip()
            continue

        # ── Bill header row ───────────────────────────────────────────────
        if re.match(r'\d{6}', stripped) and "Billing Period" not in stripped:
            parts = stripped.split()
            if parts and parts[0].isdigit() and len(parts[0]) == 6:
                current["bill_id"] = parts[0]
                cost_m = re.search(r'(\d[\d,]*\.\d{4})\s*$', line.strip())
                if cost_m:
                    current["cost"] = float(cost_m.group(1).replace(",", ""))
                pm = re.search(r'\t(20\d{4})\t', line)
                if pm:
                    current["billing_period"] = pm.group(1)
            continue

        # ── Flag Type line (status, assignee, cost recovery) ─────────────
        if "Flag Type:" in stripped:
            sm = re.search(r'Flag Status:\s*(\w+)', stripped)
            if sm:
                current["status"] = sm.group(1)

            am = re.search(r'Assigned to:\s*\t+(.+?)(?:\t{4,}|Cost Recovery)', stripped)
            if am:
                current["assigned_to"] = am.group(1).strip()

            rm = re.search(r'Cost Recovery:\s*\$([0-9,.]+)', stripped)
            if rm:
                current["cost_recovery"] = float(rm.group(1).replace(",", ""))
            continue

        # ── Flag Issue line (separate row from Flag Type) ─────────────────
        if stripped.startswith("Flag Issue:"):
            im = re.search(r'Flag Issue:\s*(.+?)(?:\t{2,}|$)', stripped)
            if im:
                issues_raw = im.group(1).strip()
                current["flag_issues"] = issues_raw
                current["num_issues"] = len([x for x in issues_raw.split(",") if x.strip()])
            continue

        # ── Timeline events ───────────────────────────────────────────────
        if re.match(r'\d{2}/\d{2}/\d{4}', stripped):
            dt_m = re.match(r'(\d{2}/\d{2}/\d{4} \d{2}:\d{2} (?:AM|PM))', stripped)
            if dt_m:
                try:
                    dt = datetime.strptime(dt_m.group(1), "%m/%d/%Y %I:%M %p")
                except ValueError:
                    dt = None

                if "Bill flagged" in stripped or "flagged as Audit" in stripped:
                    if current.get("flagged_date") is None and dt:
                        current["flagged_date"] = dt
                    current["flag_events"].append({"type": "flagged", "dt": dt, "actor": "SYSTEM"})

                elif "Flag resolved" in stripped and dt:
                    current["resolved_date"] = dt
                    actor_m = re.search(r'\d{2}:\d{2} (?:AM|PM) ([\w.@]+) Flag resolved', stripped)
                    actor = actor_m.group(1) if actor_m else "SYSTEM"
                    current["resolvers"].append(actor)
                    current["flag_events"].append({"type": "resolved", "dt": dt, "actor": actor})

    if current.get("bill_id"):
        records.append(current)

    if not records:
        return pd.DataFrame()

    df = pd.DataFrame(records)

    # ── Derived columns ───────────────────────────────────────────────────
    df["billing_period_dt"] = pd.to_datetime(df["billing_period"], format="%Y%m", errors="coerce")
    df["billing_period_label"] = df["billing_period_dt"].dt.strftime("%b %Y")

    df["days_to_resolve"] = df.apply(
        lambda r: (r["resolved_date"] - r["flagged_date"]).days
        if r["resolved_date"] and r["flagged_date"] else None, axis=1
    )

    df["primary_resolver"] = df["resolvers"].apply(lambda x: x[-1] if x else "Unresolved")

    df["issues_list"] = df["flag_issues"].apply(
        lambda x: [i.strip() for i in x.split(",") if i.strip()] if x else []
    )

    df["primary_issue"] = df["issues_list"].apply(lambda x: x[0] if x else "Unknown")
    df["primary_category"] = df["primary_issue"].apply(
        lambda x: FLAG_META.get(x, {}).get("category", "Other")
    )
    df["primary_priority"] = df["primary_issue"].apply(
        lambda x: FLAG_META.get(x, {}).get("priority", "Medium")
    )

    return df


def _xlsx_to_text_openpyxl(path: str) -> str:
    """Fallback: convert xlsx to tab-separated text using openpyxl."""
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)
    lines = []
    for ws in wb.worksheets:
        lines.append(f"## Sheet: {ws.title}")
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            lines.append("\t" + "\t".join(cells))
    return "\n".join(lines)


def load_uploaded_file(uploaded_file) -> pd.DataFrame:
    """
    Accept a Streamlit UploadedFile (.xlsx), parse it and return a DataFrame.
    Uses extract-text CLI if available, otherwise falls back to openpyxl.
    """
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp.write(uploaded_file.read())
        tmp_path = tmp.name

    text = ""
    try:
        if shutil.which("extract-text"):
            result = subprocess.run(
                ["extract-text", tmp_path],
                capture_output=True, text=True, timeout=60,
            )
            text = result.stdout
        if not text.strip():
            text = _xlsx_to_text_openpyxl(tmp_path)
    finally:
        os.unlink(tmp_path)

    return parse_report27_text(text) if text.strip() else pd.DataFrame()
