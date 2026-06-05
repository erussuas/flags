"""
ec_parser.py — EnergyCAP Report-27, Report-18, Report-03, Report-19 parsers
                + emissions classification + configuration health checks
                + continuity audit logic
"""
import re, os, shutil, subprocess, tempfile
from datetime import datetime
from collections import Counter, defaultdict

import pandas as pd
from openpyxl import load_workbook


# ══════════════════════════════════════════════════════════════════════════════
# FLAG METADATA  (Report-27)
# ══════════════════════════════════════════════════════════════════════════════
FLAG_META = {
    "Rate schedule mismatch":              {"category":"Import / Configuration","priority":"Medium","cause":"Rate schedule in import file doesn't match the rate schedule assigned to the meter in EnergyCAP.","action":"Update the meter's rate schedule in EnergyCAP to match the utility's current rate, or correct the import file.","in_energycap":"Meter record → Rate Schedule field."},
    "Serial number mismatch":              {"category":"Import / Configuration","priority":"Medium","cause":"Serial number in the import file doesn't match EnergyCAP. May indicate the vendor swapped the physical meter.","action":"Verify with the vendor whether the meter was replaced. If replaced, update the serial number in EnergyCAP.","in_energycap":"Meter record → Serial Number field. If meter was swapped, retire old meter and create a new one."},
    "Conflicting use units":               {"category":"Import / Configuration","priority":"Medium","cause":"The use unit on the bill conflicts with the use unit on the meter in EnergyCAP.","action":"Determine the correct unit of measure. Update the meter configuration in EnergyCAP or correct the import file.","in_energycap":"Meter record → Use Unit field."},
    "Duplicate bill":                      {"category":"Duplicate / Overlap","priority":"High","cause":"Total bill cost equals a prior bill's cost AND start/end dates match — likely a data entry or billing error.","action":"Compare with the prior bill. If truly duplicate, void one.","in_energycap":"More Actions → Void Bill."},
    "Overlapping bill":                    {"category":"Duplicate / Overlap","priority":"High","cause":"One or more bills have overlapping start/end dates on the same account.","action":"Review dates of all overlapping bills. Correct dates if data entry error.","in_energycap":"Account History tab → view adjacent bills."},
    "Multiple bills in period":            {"category":"Duplicate / Overlap","priority":"High","cause":"More than one bill exists for the same account and billing period.","action":"Determine if split bills (valid) or accidental duplicates. Void true duplicates.","in_energycap":"Filter bill list by account and period. More Actions → Void Bill."},
    "Abnormal cost":                       {"category":"Statistical Outlier","priority":"High","cause":"Total cost is a severe statistical outlier (>3.0 std dev) vs last 12 bills using quadratic regression.","action":"Review bill line items for unexpected charges. Check if a rate change occurred. Consider cost recovery if error found.","in_energycap":"Open bill → check line items. Account History tab to compare to prior bills."},
    "Abnormal use":                        {"category":"Statistical Outlier","priority":"High","cause":"Usage is a severe statistical outlier compared to the last 12 bills. Could indicate meter issue, leak, or billing error.","action":"Check for operational changes at the site. Contact vendor if usage spike is unexplained.","in_energycap":"Open bill → review meter use values. Compare to Account History."},
    "Abnormal demand":                     {"category":"Statistical Outlier","priority":"High","cause":"Demand reading is a severe statistical outlier compared to historical bills.","action":"Investigate if new high-load equipment was added. Verify demand meter readings with the vendor.","in_energycap":"Review demand meter readings via Account History."},
    "Gap between bills":                   {"category":"Date / Timeline","priority":"Medium","cause":"Gap of 2+ days between this bill and the preceding bill — could mean a missing bill.","action":"Check if a bill was skipped or not yet received. Contact vendor or check vendor portal.","in_energycap":"Account History tab to see the gap. If a bill is missing, manually enter it or re-import."},
    "Shorter or longer bill":              {"category":"Date / Timeline","priority":"Low","cause":"Bill period is significantly shorter or longer than the average of prior bills.","action":"Verify with the vendor that the billing period is correct. If dates are off, correct them.","in_energycap":"Review Start and End Date on the bill header."},
    "Late statement date":                 {"category":"Date / Timeline","priority":"Low","cause":"Statement date is too many days after the bill's end date.","action":"Verify the statement date is correct.","in_energycap":"Correct the Statement Date in the bill header."},
    "Late due date":                       {"category":"Date / Timeline","priority":"Low","cause":"Due date is too many days after the bill's end date.","action":"Verify due date accuracy to avoid late payment penalties.","in_energycap":"Correct the Due Date in the bill header."},
    "Unexpected billing period":           {"category":"Date / Timeline","priority":"Medium","cause":"Billing period is outside the bill's start and end dates — likely a data entry error.","action":"Correct the billing period to align with the start and end dates of the bill.","in_energycap":"Edit the bill header → set Billing Period to match start/end dates."},
    "Flagged line item type found":        {"category":"Line Item Review","priority":"Medium","cause":"Bill contains a line item type configured for monitoring (e.g., late fees, taxes, penalties).","action":"Review the specific line item. If it's a late fee, confirm the prior bill was paid on time.","in_energycap":"Open bill → scroll to line items section → review the flagged line item type."},
    "Flagged line item description found": {"category":"Line Item Review","priority":"Medium","cause":"Bill contains a line item with a description matching a monitored keyword (case-insensitive).","action":"Review the line item description and value. Confirm whether charge is legitimate.","in_energycap":"Open bill → check line items for the matching description."},
    "High use per day":                    {"category":"Statistical Outlier","priority":"Medium","cause":"Use per day exceeds 3× the highest use-per-day from the top 90% of prior bills (last 4 years).","action":"Investigate operational changes or equipment issues at the site.","in_energycap":"Compare current use/day to historical via Account History tab."},
    "High cost per day":                   {"category":"Statistical Outlier","priority":"Medium","cause":"Cost per day exceeds 3× the highest cost-per-day from the top 90% of prior bills.","action":"Check for rate changes, new charges, or usage spikes driving up cost.","in_energycap":"Open bill and compare line-by-line with a prior period bill."},
}

PRIORITY_ORDER = {"High": 0, "Medium": 1, "Low": 2}
PRIORITY_COLOR = {"High": "#dc3545", "Medium": "#fd7e14", "Low": "#198754"}
CATEGORY_COLOR = {
    "Import / Configuration": "#0d6efd",
    "Duplicate / Overlap":    "#dc3545",
    "Statistical Outlier":    "#6f42c1",
    "Date / Timeline":        "#fd7e14",
    "Line Item Review":       "#198754",
}


# ══════════════════════════════════════════════════════════════════════════════
# ENERGY & EMISSIONS CLASSIFICATION
# ══════════════════════════════════════════════════════════════════════════════
ENERGY_COMMODITIES = {          # EnergyCAP commodity → display category
    "Electric":                 "Electricity",
    "Lighting":                 "Electricity",
    "Solar PV":                 "Electricity",
    "Natural Gas":              "Natural Gas",
    "Methane":                  "Natural Gas",
    "Biogas":                   "Natural Gas",
    "Other Gaseous Fuels":      "Natural Gas",
    "Liquified Petroleum Gases":"LPG / Propane",
    "Propane":                  "LPG / Propane",
    "Butane":                   "LPG / Propane",
    "Biomass":                  "Biomass",
    "Delivered Heat":           "District Heat / Steam",
    "Steam":                    "District Heat / Steam",
    "Diesel Fuel":              "Diesel / Fuel Oil",
    "Fuel Oil":                 "Diesel / Fuel Oil",
    "Gasoline":                 "Gasoline",
    "Aviation Fuel":            "Aviation Fuel",
    "Coal":                     "Coal",
}

GHG_SCOPE = {
    "Electricity":            "Scope 2",
    "Natural Gas":            "Scope 1",
    "LPG / Propane":          "Scope 1",
    "Biomass":                "Scope 1 (biogenic)",
    "District Heat / Steam":  "Scope 2",
    "Diesel / Fuel Oil":      "Scope 1",
    "Gasoline":               "Scope 1",
    "Aviation Fuel":          "Scope 1",
    "Coal":                   "Scope 1",
}

# Unit conversion to kWh (internal base for cross-commodity comparison)
TO_KWH = {
    "kWh":1.0,"MWh":1000.0,"GJ":277.778,"MJ":0.277778,
    "MMBtu":293.071,"THERM":29.3071,"DKTHM":293.071,"DKTH":293.071,
    "MCF":293.071,"CCF":29.3071,"CF":0.293071,"m³":10.55,
    "ekWh":1.0,               # steam already in kWh equivalent
    "t":1630.0,               # biomass tonnes → kWh (approx 5.86 GJ/t dry wood)
}

FROM_KWH = {
    "kWh":1.0,"MWh":1e-3,"GJ":3.6e-3,"MMBtu":3.41214e-3,"THERM":3.41214e-2,
}
DISPLAY_UNITS = ["kWh","MWh","GJ","MMBtu","THERM"]

# Water/non-energy fallback detection
ALWAYS_WATER_UNITS  = {"gal","Kgal","Hgal","Mgal","L"}
ENERGY_UNITS_SET    = set(TO_KWH.keys())
WATER_CAPTION_RE    = re.compile(
    r'water|sewer|sewage|irrigation|\(water\)|\(sewer\)|h2o|fire line|fire service',
    re.IGNORECASE)

def kwh(value, unit):
    return value * TO_KWH.get(unit, 0.0)

def from_kwh(kwh_val, display_unit):
    return kwh_val * FROM_KWH.get(display_unit, 1.0)

def period_sort_key(p):
    """Sort key for MM-YYYY period strings."""
    try:
        m, y = p.split('-')
        return int(y) * 100 + int(m)
    except Exception:
        return 0


# ══════════════════════════════════════════════════════════════════════════════
# REPORT-03 PARSER  (master/setup data)
# ══════════════════════════════════════════════════════════════════════════════
def parse_report03(path: str) -> pd.DataFrame:
    wb = load_workbook(path, data_only=True)
    ws = wb['Sheet1']
    rows = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[1] is None: continue
        commodity = str(row[51]).strip() if row[51] else ''
        rows.append({
            'account_name':      str(row[0]).strip()  if row[0]  else '',
            'account_number':    str(row[1]).strip()  if row[1]  else '',
            'account_status':    str(row[3]).strip()  if row[3]  else '',
            'excluded_audits':   str(row[5]).strip()  if row[5]  else '',
            'cost_center_name':  str(row[17]).strip() if row[17] else '',
            'cost_center_code':  str(row[18]).strip() if row[18] else '',
            'vendor_name':       str(row[19]).strip() if row[19] else '',
            'vendor_code':       str(row[20]).strip() if row[20] else '',
            'vendor_role':       str(row[21]).strip() if row[21] else '',
            'rate_schedule':     str(row[22]).strip() if row[22] else '',
            'meter_name':        str(row[25]).strip() if row[25] else '',
            'meter_code':        str(row[26]).strip() if row[26] else '',
            'serial_number':     str(row[27]).strip() if row[27] else '',
            'meter_status':      str(row[30]).strip() if row[30] else '',
            'acct_meter_begin':  row[31],
            'acct_meter_end':    row[32],
            'billing_frequency': str(row[35]).strip() if row[35] else '',
            'commodity':         commodity,
            'building_name':     str(row[52]).strip() if row[52] else '',
            'building_code':     str(row[53]).strip() if row[53] else '',
            'floor_area':        row[54],
            'primary_use':       str(row[57]).strip() if row[57] else '',
            'weather_station':   str(row[58]).strip() if row[58] else '',
            'building_country':  str(row[61]).strip() if row[61] else '',
            'legal_entity':      str(row[69]).strip() if row[69] else '',
            'currency_code':     str(row[82]).strip() if row[82] else '',
            'energy_category':   ENERGY_COMMODITIES.get(commodity, ''),
            'is_energy':         commodity in ENERGY_COMMODITIES,
            'ghg_scope':         GHG_SCOPE.get(ENERGY_COMMODITIES.get(commodity,''),''),
        })
    return pd.DataFrame(rows).drop_duplicates('meter_code')


# ══════════════════════════════════════════════════════════════════════════════
# CONFIGURATION HEALTH CHECKS  (runs on Report-03 DataFrame)
# ══════════════════════════════════════════════════════════════════════════════
CONFIG_CHECK_META = {
    "missing_serial_energy": {
        "title":    "Energy meters missing serial number",
        "severity": "Medium",
        "why":      "Serial number is the key EnergyCAP uses to match import file readings to the "
                    "correct meter. Missing serial numbers cause serial number mismatch flags on every "
                    "imported bill and make meter swap detection impossible.",
        "action":   "Obtain the serial number from the utility bill or physical meter and update "
                    "in EnergyCAP under the meter record.",
        "in_energycap": "Sites & Meters → Meter record → Serial Number field.",
    },
    "inactive_energy": {
        "title":    "Inactive energy meters still in portfolio",
        "severity": "Low",
        "why":      "Inactive meters may still receive bills if service was not formally terminated. "
                    "Bills on inactive meters are excluded from active-account reports, creating "
                    "silent consumption gaps in emissions calculations.",
        "action":   "Confirm whether service is truly terminated. If yes, ensure no further bills "
                    "are being processed. If no, reactivate the meter.",
        "in_energycap": "Sites & Meters → Meter record → Status field.",
    },
    "duplicate_serial": {
        "title":    "Duplicate serial numbers across multiple meters",
        "severity": "High",
        "why":      "The same serial number on two different meter records means imported bills could "
                    "be applied to the wrong meter, causing double-counting or missed consumption "
                    "in the emissions inventory.",
        "action":   "Verify the physical serial number for each meter. Correct or retire the "
                    "duplicate meter record.",
        "in_energycap": "Sites & Meters → search by serial number → compare and correct meter records.",
    },
    "deregulated_double_count_risk": {
        "title":    "Deregulated accounts with split supply/distribution vendors",
        "severity": "Medium",
        "why":      "When a utility account is deregulated, both the supplier and the distribution "
                    "company send separate bills. If both are included in emissions queries without "
                    "filtering by vendor role, consumption is double-counted.",
        "action":   "In all EnergyCAP reports and emissions queries, add the filter "
                    "'Bill is from external vendor = From external vendor' and ensure only one "
                    "vendor role per account is used for consumption reporting.",
        "in_energycap": "Report filters → Bill is from external vendor → From external vendor.",
    },
    "buildings_no_energy_meters": {
        "title":    "Buildings with no energy meters assigned",
        "severity": "Medium",
        "why":      "Buildings with active floor area but no energy meters may represent gaps in "
                    "the emissions inventory — either the meters exist in a different system, were "
                    "never set up, or the building is genuinely non-metered.",
        "action":   "Review each building. If it has utility consumption, ensure the relevant "
                    "meters are created and linked in EnergyCAP.",
        "in_energycap": "Sites & Meters → Building record → Associated meters.",
    },
    "missing_floor_area_energy": {
        "title":    "Energy meters on buildings missing floor area",
        "severity": "Low",
        "why":      "Floor area is required to calculate energy use intensity (EUI) and emissions "
                    "intensity per square foot — key normalisation metrics for benchmarking and "
                    "target-setting.",
        "action":   "Obtain the gross floor area (GFA) from property records or facility management "
                    "and update in EnergyCAP.",
        "in_energycap": "Sites & Meters → Building record → Floor Area field.",
    },
    "billing_freq_typos": {
        "title":    "Billing frequency field contains non-standard values",
        "severity": "Low",
        "why":      "Non-standard values indicate data entry errors. Billing frequency is used for "
                    "gap detection and accrual calculations — incorrect values cause false gap flags "
                    "and inaccurate accruals.",
        "action":   "Correct to standard values: Monthly, Quarterly, Annual, Half Yearly, "
                    "BiMonthly, Intermittent.",
        "in_energycap": "Sites & Meters → Meter record → Billing Frequency field.",
    },
    "non_usd_currency": {
        "title":    "Non-USD currency accounts in portfolio",
        "severity": "Low",
        "why":      "Bills in non-USD currencies require exchange rate conversion for cost reporting. "
                    "If exchange rates are not maintained in EnergyCAP, cost-based emissions "
                    "allocations and budget variances will be incorrect.",
        "action":   "Verify exchange rates are configured and kept current for all non-USD accounts. "
                    "Document the exchange rate methodology used for emissions reporting.",
        "in_energycap": "Administration → Currency Exchange Rates.",
    },
    "excluded_from_audits": {
        "title":    "Energy meters excluded from bill audits",
        "severity": "Medium",
        "why":      "Meters excluded from audits will not generate flags even if bills contain "
                    "errors — abnormal usage, duplicate bills, and rate mismatches all pass through "
                    "silently. This creates blind spots in the emissions data quality process.",
        "action":   "Review whether exclusions are still justified. Re-enable audits for any meter "
                    "where the original reason for exclusion no longer applies.",
        "in_energycap": "Sites & Meters → Meter record → Excluded From Audits field.",
    },
}

SEVERITY_ORDER = {"High": 0, "Medium": 1, "Low": 2}

def _is_blank(s):
    return str(s).strip() in {'', 'None', 'nan', 'NaN', 'none'}

def run_config_checks(df: pd.DataFrame) -> dict:
    results = {}
    energy = df[df['is_energy']].copy()

    # 1. Energy meters missing serial number
    rows = energy[energy['serial_number'].apply(_is_blank)]
    results['missing_serial_energy'] = {
        **CONFIG_CHECK_META['missing_serial_energy'],
        'count': len(rows), 'rows': rows,
    }

    # 2. Inactive energy meters
    rows = energy[energy['meter_status'] == 'Inactive']
    results['inactive_energy'] = {
        **CONFIG_CHECK_META['inactive_energy'],
        'count': len(rows), 'rows': rows,
    }

    # 3. Duplicate serial numbers
    valid = df[~df['serial_number'].apply(_is_blank)]
    dup = (valid.groupby('serial_number')['meter_code']
           .apply(list).reset_index())
    dup = dup[dup['meter_code'].apply(len) > 1]
    dup_meters = valid[valid['serial_number'].isin(dup['serial_number'])]
    results['duplicate_serial'] = {
        **CONFIG_CHECK_META['duplicate_serial'],
        'count': len(dup), 'rows': dup_meters, 'detail': dup,
    }

    # 4. Deregulated double-count risk
    rows = df[df['vendor_role'].isin(['Supply', 'Distribution'])]
    results['deregulated_double_count_risk'] = {
        **CONFIG_CHECK_META['deregulated_double_count_risk'],
        'count': len(rows), 'rows': rows,
    }

    # 5. Buildings with no energy meters
    all_bldg = set(df[~df['building_code'].apply(_is_blank)]['building_code'].unique())
    nrg_bldg = set(energy[~energy['building_code'].apply(_is_blank)]['building_code'].unique())
    no_nrg   = all_bldg - nrg_bldg
    rows = df[df['building_code'].isin(no_nrg)].drop_duplicates('building_code')
    results['buildings_no_energy_meters'] = {
        **CONFIG_CHECK_META['buildings_no_energy_meters'],
        'count': len(no_nrg), 'rows': rows,
    }

    # 6. Energy meters missing floor area
    rows = energy[energy['floor_area'].isnull() | (energy['floor_area'] == 0)]
    results['missing_floor_area_energy'] = {
        **CONFIG_CHECK_META['missing_floor_area_energy'],
        'count': len(rows), 'rows': rows,
    }

    # 7. Billing frequency typos
    standard = {'Monthly','Quarterly','Annual','Half Yearly','BiMonthly',
                'Intermittent','','None','nan'}
    rows = df[~df['billing_frequency'].isin(standard)]
    results['billing_freq_typos'] = {
        **CONFIG_CHECK_META['billing_freq_typos'],
        'count': len(rows), 'rows': rows,
    }

    # 8. Non-USD currency
    rows = df[~df['currency_code'].apply(_is_blank) & ~df['currency_code'].isin(['USD'])]
    results['non_usd_currency'] = {
        **CONFIG_CHECK_META['non_usd_currency'],
        'count': len(rows), 'rows': rows,
    }

    # 9. Energy meters excluded from audits
    rows = energy[energy['excluded_audits'] == 'Yes']
    results['excluded_from_audits'] = {
        **CONFIG_CHECK_META['excluded_from_audits'],
        'count': len(rows), 'rows': rows,
    }

    return results


# ══════════════════════════════════════════════════════════════════════════════
# REPORT-19 PARSER  (monthly use and cost — continuity audit)
# ══════════════════════════════════════════════════════════════════════════════
def parse_report19(path: str, setup_df: pd.DataFrame = None) -> pd.DataFrame:
    """
    Parse Report-19 Monthly Utility Use and Cost (data-only Excel export).
    One sheet per meter. Returns a long-format DataFrame:
      meter_name, commodity, serial_number, use_unit, energy_category,
      period, use, cost, demand, kwh_equivalent, is_energy
    """
    wb = load_workbook(path, data_only=True)
    sheets = [s for s in wb.sheetnames if s.lower() != 'report overview']

    # Build serial → energy_category lookup from setup if available
    serial_lookup = {}
    meter_lookup  = {}
    if setup_df is not None and not setup_df.empty:
        for _, row in setup_df.iterrows():
            if row['serial_number'] and not _is_blank(row['serial_number']):
                serial_lookup[str(row['serial_number']).strip()] = {
                    'energy_category': row['energy_category'],
                    'is_energy':       row['is_energy'],
                    'ghg_scope':       row['ghg_scope'],
                    'building_code':   row['building_code'],
                    'building_name':   row['building_name'],
                    'meter_code':      row['meter_code'],
                    'billing_frequency': row['billing_frequency'],
                }
            if row['meter_code'] and not _is_blank(row['meter_code']):
                meter_lookup[str(row['meter_code']).strip()] = serial_lookup.get(
                    str(row['serial_number']).strip(), {})

    all_rows = []

    for sheet_name in sheets:
        ws   = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 9:
            continue

        try:
            meter_name = str(rows[1][0]).strip() if rows[1][0] else sheet_name
            commodity  = str(rows[2][0]).strip() if rows[2][0] else ''
            serial_num = str(rows[4][0]).strip() if rows[4][0] else ''

            # Detect with/without demand column from header row (row index 7)
            header_vals = [str(v).strip() if v is not None else '' for v in rows[7]]
            non_empty   = [v for v in header_vals if v]
            has_demand  = any('demand' in v.lower() or v.lower().endswith('kw')
                              for v in non_empty)

            # Extract use unit from "Use kWh", "Use DKTHM", etc.
            use_unit_raw = next((v for v in non_empty if v.lower().startswith('use ')), '')
            use_unit = use_unit_raw[4:].strip() if use_unit_raw else ''

            use_col  = 3
            cost_col = 5 if has_demand else 4
            dem_col  = 4 if has_demand else None

            # Classify commodity
            energy_cat = ENERGY_COMMODITIES.get(commodity, '')
            is_energy  = commodity in ENERGY_COMMODITIES
            scope      = GHG_SCOPE.get(energy_cat, '')

            # Enrich from setup if serial matches
            setup_meta = serial_lookup.get(serial_num, {})
            if not is_energy and setup_meta.get('is_energy'):
                is_energy  = setup_meta['is_energy']
                energy_cat = setup_meta.get('energy_category', energy_cat)
                scope      = setup_meta.get('ghg_scope', scope)

            if not is_energy:
                continue  # skip non-energy meters entirely

            # Parse data rows
            for row in rows[8:]:
                if row[0] is None: continue
                period = str(row[0]).strip()
                if period.lower() == 'total': break
                if not re.match(r'[0-9]{2}-[0-9]{4}', period): continue

                use_val = row[use_col]
                cost_val = row[cost_col]
                dem_val  = row[dem_col] if dem_col else None

                try:    use  = float(use_val)  if use_val  is not None else None
                except: use  = None
                try:    cost = float(cost_val) if cost_val is not None else None
                except: cost = None
                try:    dem  = float(dem_val)  if dem_val  is not None else None
                except: dem  = None

                kwh_eq = kwh(use, use_unit) if use is not None else None

                all_rows.append({
                    'sheet':          sheet_name,
                    'meter_name':     meter_name,
                    'serial_number':  serial_num,
                    'commodity':      commodity,
                    'energy_category':energy_cat,
                    'ghg_scope':      scope,
                    'use_unit':       use_unit,
                    'period':         period,
                    'period_sort':    period_sort_key(period),
                    'use':            use,
                    'cost':           cost,
                    'demand':         dem,
                    'kwh_equivalent': kwh_eq,
                    'building_code':  setup_meta.get('building_code',''),
                    'building_name':  setup_meta.get('building_name',''),
                    'meter_code':     setup_meta.get('meter_code',''),
                    'billing_frequency': setup_meta.get('billing_frequency',''),
                })

        except Exception:
            continue

    if not all_rows:
        return pd.DataFrame()

    return pd.DataFrame(all_rows)


# ══════════════════════════════════════════════════════════════════════════════
# CONTINUITY AUDIT  (runs on parsed Report-19 DataFrame)
# ══════════════════════════════════════════════════════════════════════════════
CONTINUITY_CHECK_META = {
    "large_gaps": {
        "title":    "Large billing gaps (3+ consecutive months missing)",
        "severity": "High",
        "why":      "Three or more consecutive missing months on an active energy meter almost "
                    "certainly represents missing consumption data in the emissions inventory. "
                    "Unlike a single skipped bill (which could be timing), a multi-month gap "
                    "is a confirmed understatement of Scope 1 or Scope 2 emissions.",
        "action":   "Contact the vendor to obtain back-bills for the missing periods. "
                    "If bills were received but not entered, import them. Check whether the "
                    "meter was on a seasonal or temporary hold.",
        "in_energycap": "Bills → Account History for the affected account → identify missing periods.",
    },
    "medium_gaps": {
        "title":    "Medium billing gaps (2 consecutive months missing)",
        "severity": "Medium",
        "why":      "Two consecutive missing months may indicate a missing bill, a billing "
                    "cycle change, or a quarterly-billed meter where only some quarters "
                    "were imported. Cross-reference with billing frequency in the setup data.",
        "action":   "Check the meter's billing frequency in EnergyCAP. If monthly, investigate "
                    "missing bills. If quarterly, confirm the cadence is correctly configured.",
        "in_energycap": "Bills → Account History. Meter record → Billing Frequency field.",
    },
    "small_gaps": {
        "title":    "Single-month billing gaps",
        "severity": "Low",
        "why":      "A single missing month could be a delayed bill, a billing cycle shift, "
                    "or a genuine missing import. Less urgent than multi-month gaps but "
                    "should be reviewed for monthly-billed energy meters.",
        "action":   "Verify whether the bill was received and not yet entered, or whether "
                    "the billing cycle shifted. Check the vendor portal.",
        "in_energycap": "Bills → Account History. Check for bills in pending batches.",
    },
    "zero_use_nonzero_cost": {
        "title":    "Zero consumption with non-zero cost",
        "severity": "Medium",
        "why":      "A period showing zero energy use but a positive charge is either a "
                    "fixed demand/capacity charge on an idle account, a data entry error "
                    "where use was not recorded, or an account that should have been closed. "
                    "These periods contribute zero emissions but may represent real consumption "
                    "that was not captured.",
        "action":   "Review each instance. If the account was idle (vacant facility, standby "
                    "service), document it. If use data is missing, obtain it from the vendor. "
                    "If the account should be closed, initiate closure.",
        "in_energycap": "Open the bill → check if use line items were imported correctly.",
    },
    "negative_use": {
        "title":    "Negative energy consumption values",
        "severity": "High",
        "why":      "Negative use values typically indicate solar/renewable export credits, "
                    "bill reversals, or data entry errors. If treated as real consumption, "
                    "they net against other usage and understate the emissions inventory.",
        "action":   "Identify the source of each negative value. If it is a solar export "
                    "credit, ensure it is correctly classified and handled in your emissions "
                    "accounting methodology. If it is a reversal, verify that the corrected "
                    "bill was also entered.",
        "in_energycap": "Open the bill → review line items for credits or reversals.",
    },
    "new_meters": {
        "title":    "Energy meters with data only in the most recent year",
        "severity": "Low",
        "why":      "Meters appearing for the first time in the reporting period may be "
                    "genuinely new installations, or may be pre-existing meters that were "
                    "recently added to EnergyCAP. In either case, verify that the setup "
                    "configuration is complete and that prior-period data was backfilled "
                    "if required for the emissions baseline.",
        "action":   "Confirm the meter's installation date. If the meter was installed "
                    "before the reporting period, determine whether historical bills are "
                    "available for backfill.",
        "in_energycap": "Meter record → Acct-Meter Begin Date. Check for prior bills.",
    },
    "single_year_only": {
        "title":    "Energy meters with data only in the prior year (no recent data)",
        "severity": "Medium",
        "why":      "A meter with data in the prior year but nothing in the current year "
                    "may represent a decommissioned account, a closed facility, or a meter "
                    "that stopped being imported. Any of these could mean missing current-year "
                    "emissions data.",
        "action":   "Verify whether the account is still active. If service was terminated, "
                    "confirm the termination date and close the account in EnergyCAP. "
                    "If service continues, investigate why recent bills are absent.",
        "in_energycap": "Account record → Service Dates. Bills → Account History.",
    },
}

def run_continuity_checks(df19: pd.DataFrame) -> dict:
    """
    Run continuity audit on parsed Report-19 DataFrame.
    Returns dict of check_key → {meta, count, rows DataFrame}
    """
    if df19.empty:
        return {k: {**v, 'count': 0, 'rows': pd.DataFrame()}
                for k, v in CONTINUITY_CHECK_META.items()}

    results = {}

    # Determine year 1 and year 2 period sets from the data
    all_periods = sorted(df19['period'].unique(), key=period_sort_key)
    n = len(all_periods)
    mid = n // 2
    yr1_periods = set(all_periods[:mid]) if n > 1 else set()
    yr2_periods = set(all_periods[mid:]) if n > 1 else set(all_periods)

    # Per-meter analysis
    gap_rows_large, gap_rows_medium, gap_rows_small = [], [], []
    zero_cost_rows   = []
    neg_use_rows     = []
    new_meter_rows   = []
    gone_meter_rows  = []

    for sheet, grp in df19.groupby('sheet'):
        meta_row = grp.iloc[0]
        meter    = meta_row['meter_name']
        commodity= meta_row['commodity']
        cat      = meta_row['energy_category']
        scope    = meta_row['ghg_scope']
        bfreq    = meta_row.get('billing_frequency', '')
        bldg     = meta_row.get('building_name', '')

        meter_periods = sorted(grp['period'].unique(), key=period_sort_key)
        period_set    = set(meter_periods)

        has_yr1 = bool(period_set & yr1_periods)
        has_yr2 = bool(period_set & yr2_periods)

        # New meter (only yr2)
        if has_yr2 and not has_yr1 and yr1_periods:
            first_period = meter_periods[0]
            new_meter_rows.append({
                'meter_name': meter, 'commodity': commodity,
                'energy_category': cat, 'ghg_scope': scope,
                'building_name': bldg, 'first_period': first_period,
                'n_periods': len(meter_periods),
                'billing_frequency': bfreq,
            })

        # Gone meter (only yr1)
        if has_yr1 and not has_yr2 and yr2_periods:
            last_period = meter_periods[-1]
            gone_meter_rows.append({
                'meter_name': meter, 'commodity': commodity,
                'energy_category': cat, 'ghg_scope': scope,
                'building_name': bldg, 'last_period': last_period,
                'n_periods': len(meter_periods),
                'billing_frequency': bfreq,
            })

        # Gap detection within meter periods
        for i in range(len(meter_periods) - 1):
            m1, y1 = meter_periods[i].split('-')
            m2, y2 = meter_periods[i+1].split('-')
            apart  = (int(y2) - int(y1)) * 12 + (int(m2) - int(m1))
            if apart <= 1:
                continue
            missing = apart - 1
            gap_info = {
                'meter_name':     meter,
                'commodity':      commodity,
                'energy_category':cat,
                'ghg_scope':      scope,
                'building_name':  bldg,
                'gap_after':      meter_periods[i],
                'gap_before':     meter_periods[i+1],
                'months_missing': missing,
                'billing_frequency': bfreq,
            }
            if missing >= 3:
                gap_rows_large.append(gap_info)
            elif missing == 2:
                gap_rows_medium.append(gap_info)
            else:
                gap_rows_small.append(gap_info)

        # Zero use / non-zero cost
        for _, row in grp.iterrows():
            if (row['use'] is not None and row['use'] == 0
                    and row['cost'] is not None and row['cost'] > 0):
                zero_cost_rows.append({
                    'meter_name':      meter,
                    'commodity':       commodity,
                    'energy_category': cat,
                    'ghg_scope':       scope,
                    'building_name':   bldg,
                    'period':          row['period'],
                    'cost':            row['cost'],
                })

            # Negative use
            if row['use'] is not None and row['use'] < 0:
                neg_use_rows.append({
                    'meter_name':      meter,
                    'commodity':       commodity,
                    'energy_category': cat,
                    'ghg_scope':       scope,
                    'building_name':   bldg,
                    'period':          row['period'],
                    'use':             row['use'],
                    'use_unit':        row['use_unit'],
                    'cost':            row['cost'],
                })

    def to_df(rows): return pd.DataFrame(rows) if rows else pd.DataFrame()

    results['large_gaps']           = {**CONTINUITY_CHECK_META['large_gaps'],
                                        'count': len(set(r['meter_name'] for r in gap_rows_large)),
                                        'rows': to_df(gap_rows_large)}
    results['medium_gaps']          = {**CONTINUITY_CHECK_META['medium_gaps'],
                                        'count': len(set(r['meter_name'] for r in gap_rows_medium)),
                                        'rows': to_df(gap_rows_medium)}
    results['small_gaps']           = {**CONTINUITY_CHECK_META['small_gaps'],
                                        'count': len(set(r['meter_name'] for r in gap_rows_small)),
                                        'rows': to_df(gap_rows_small)}
    results['zero_use_nonzero_cost'] = {**CONTINUITY_CHECK_META['zero_use_nonzero_cost'],
                                        'count': len(zero_cost_rows),
                                        'rows': to_df(zero_cost_rows)}
    results['negative_use']          = {**CONTINUITY_CHECK_META['negative_use'],
                                        'count': len(neg_use_rows),
                                        'rows': to_df(neg_use_rows)}
    results['new_meters']            = {**CONTINUITY_CHECK_META['new_meters'],
                                        'count': len(new_meter_rows),
                                        'rows': to_df(new_meter_rows)}
    results['single_year_only']      = {**CONTINUITY_CHECK_META['single_year_only'],
                                        'count': len(gone_meter_rows),
                                        'rows': to_df(gone_meter_rows)}
    return results


# ══════════════════════════════════════════════════════════════════════════════
# REPORT-18 PARSER  (bill line item usage — emissions priority)
# ══════════════════════════════════════════════════════════════════════════════
def parse_report18(path: str, setup_df=None) -> pd.DataFrame:
    wb = load_workbook(path, data_only=True)
    meter_lookup = {}
    if setup_df is not None and not setup_df.empty:
        meter_lookup = setup_df.set_index('meter_code')[
            ['commodity','energy_category','is_energy','ghg_scope']
        ].to_dict('index')

    all_rows = []
    for sheet_name in wb.sheetnames:
        if sheet_name.lower() in {'report overview','overview','demand','info_use'}:
            continue
        ws = wb[sheet_name]
        header_idx  = {}
        data_started = False
        for row in ws.iter_rows(values_only=True):
            vals     = list(row)
            str_vals = [str(v).strip() if v is not None else '' for v in vals]
            if 'Bill ID' in str_vals and 'Unit' in str_vals and not data_started:
                for j, v in enumerate(str_vals):
                    if v: header_idx[v] = j
                data_started = True
                continue
            if not data_started: continue
            bill_id_raw = vals[header_idx.get('Bill ID', 9)]
            if bill_id_raw is None: continue
            try:    bill_id = str(int(float(str(bill_id_raw)))).strip()
            except: continue
            meter_code = str(vals[header_idx.get('Meter Code', 1)] or '').strip()
            caption    = str(vals[header_idx.get('Caption', 12)] or '').strip()
            unit       = str(vals[header_idx.get('Unit', 14)] or '').strip()
            bp         = vals[header_idx.get('Billing Period', 8)]
            acct_code  = str(vals[header_idx.get('Account Code', 0)] or '').strip()
            try:    value = float(vals[header_idx.get('Value', 13)] or 0)
            except: value = 0.0
            if value <= 0: continue

            if meter_code in meter_lookup:
                meta        = meter_lookup[meter_code]
                is_energy   = meta['is_energy']
                energy_cat  = meta['energy_category']
                scope       = meta['ghg_scope']
            else:
                is_energy  = (unit not in ALWAYS_WATER_UNITS
                              and unit in ENERGY_UNITS_SET
                              and not WATER_CAPTION_RE.search(caption))
                energy_cat = scope = ''

            if not is_energy: continue
            kwh_eq = kwh(value, unit)
            if kwh_eq == 0: continue

            all_rows.append({
                'bill_id':        bill_id,
                'meter_code':     meter_code,
                'account_code':   acct_code,
                'billing_period': str(bp) if bp else '',
                'caption':        caption,
                'value':          value,
                'unit':           unit,
                'energy_category':energy_cat,
                'ghg_scope':      scope,
                'kwh_equivalent': kwh_eq,
            })

    if not all_rows:
        return pd.DataFrame()

    df = pd.DataFrame(all_rows)
    total_kwh = df.groupby('bill_id')['kwh_equivalent'].sum().rename('total_kwh_equivalent')
    primary = (df.sort_values('kwh_equivalent', ascending=False)
                 .drop_duplicates('bill_id')
                 [['bill_id','unit','value','kwh_equivalent',
                   'energy_category','ghg_scope','meter_code']]
                 .rename(columns={'unit':'primary_unit','value':'primary_value',
                                  'kwh_equivalent':'primary_kwh'}))
    return primary.merge(total_kwh, on='bill_id', how='left')


# ══════════════════════════════════════════════════════════════════════════════
# REPORT-27 PARSER  (bill flags)
# ══════════════════════════════════════════════════════════════════════════════
def _blank_record() -> dict:
    return {
        "account":"","address":"","vendor":"","bill_id":"",
        "billing_period":"","cost":0.0,"status":"",
        "flag_issues":"","assigned_to":"","cost_recovery":0.0,
        "flagged_date":None,"resolved_date":None,
        "num_issues":0,"resolvers":[],"flag_events":[],
    }

def parse_report27_text(text: str) -> pd.DataFrame:
    lines    = text.split("\n")
    records  = []
    current  = _blank_record()
    in_record = False

    for line in lines:
        stripped = line.strip()

        if re.search(r'Account:\s{2,}', stripped):
            if in_record and current.get("bill_id"):
                records.append(current)
            current = _blank_record()
            current["account"] = re.sub(r'Account:\s+', '', stripped).strip()
            in_record = True
            continue

        if not in_record: continue

        if (current["account"] and not current["address"]
                and not current["vendor"] and stripped
                and "Vendor:" not in stripped
                and not re.match(r'[0-9]{6}', stripped)):
            current["address"] = stripped; continue

        if "Vendor:" in stripped:
            vm = re.search(r'Vendor:\s+(.+?)(?:\s*\[|$)', stripped)
            if vm: current["vendor"] = vm.group(1).strip()
            continue

        if re.match(r'[0-9]{6}', stripped) and "Billing Period" not in stripped:
            parts = stripped.split()
            if parts and parts[0].isdigit() and len(parts[0]) == 6:
                current["bill_id"] = parts[0]
                cm = re.search(r'([0-9][\d,]*\.[0-9]+)\s*$', line.strip())
                if cm:
                    try: current["cost"] = float(cm.group(1).replace(",", ""))
                    except: pass
                pm = re.search(r'\t(20[0-9]{4})\t', line)
                if pm: current["billing_period"] = pm.group(1)
            continue

        if "Flag Type:" in stripped:
            sm = re.search(r'Flag Status:\s*(\w+)', stripped)
            if sm: current["status"] = sm.group(1)
            am = re.search(r'Assigned to:\s*\t+(.+?)(?:\t{4,}|Cost Recovery)', stripped)
            if am: current["assigned_to"] = am.group(1).strip()
            rm = re.search(r'Cost Recovery:\s*\$([0-9,.]+)', stripped)
            if rm:
                try: current["cost_recovery"] = float(rm.group(1).replace(",", ""))
                except: pass
            continue

        if stripped.startswith("Flag Issue:"):
            im = re.search(r'Flag Issue:\s*(.+?)(?:\t{2,}|$)', stripped)
            if im:
                raw = im.group(1).strip()
                current["flag_issues"] = raw
                current["num_issues"]  = len([x for x in raw.split(",") if x.strip()])
            continue

        if re.match(r'[0-9]{2}/[0-9]{2}/[0-9]{4}', stripped):
            dtm = re.match(r'([0-9]{2}/[0-9]{2}/[0-9]{4} [0-9]{2}:[0-9]{2} (?:AM|PM))', stripped)
            if dtm:
                try:    dt = datetime.strptime(dtm.group(1), "%m/%d/%Y %I:%M %p")
                except: dt = None
                if "Bill flagged" in stripped or "flagged as Audit" in stripped:
                    if not current["flagged_date"] and dt:
                        current["flagged_date"] = dt
                elif "Flag resolved" in stripped and dt:
                    current["resolved_date"] = dt
                    am2 = re.search(
                        r'[0-9]{2}:[0-9]{2} (?:AM|PM) ([\w.@]+) Flag resolved', stripped)
                    current["resolvers"].append(am2.group(1) if am2 else "SYSTEM")

    if current.get("bill_id"):
        records.append(current)
    if not records:
        return pd.DataFrame()

    df = pd.DataFrame(records)
    df["billing_period_dt"]    = pd.to_datetime(df["billing_period"], format="%Y%m", errors="coerce")
    df["billing_period_label"] = df["billing_period_dt"].dt.strftime("%b %Y")
    df["days_to_resolve"] = df.apply(
        lambda r: (r["resolved_date"] - r["flagged_date"]).days
        if r["resolved_date"] and r["flagged_date"] else None, axis=1)
    df["primary_resolver"] = df["resolvers"].apply(lambda x: x[-1] if x else "Unresolved")
    df["issues_list"] = df["flag_issues"].apply(
        lambda x: [i.strip() for i in x.split(",") if i.strip()] if x else [])
    df["primary_issue"]     = df["issues_list"].apply(lambda x: x[0] if x else "Unknown")
    df["primary_category"]  = df["primary_issue"].apply(
        lambda x: FLAG_META.get(x, {}).get("category", "Other"))
    df["primary_priority"]  = df["primary_issue"].apply(
        lambda x: FLAG_META.get(x, {}).get("priority", "Medium"))
    return df


# ══════════════════════════════════════════════════════════════════════════════
# CUSTOM FLAT FLAG FILE PARSER
# Handles the flat tabular format (one row per issue) as an alternative to R27.
# Produces the same normalized schema as parse_report27_text so the rest of
# the app works identically regardless of which format was uploaded.
# ══════════════════════════════════════════════════════════════════════════════

# Column name aliases — handles case/spacing variations across exports
_COL_ALIASES = {
    'bill_id':              ['billid','bill_id','bill id','BillID'],
    'account':              ['account','account_number','accountnumber'],
    'billing_period':       ['billingperiod','billing_period','billing period'],
    'bill_start':           ['billstart','bill_start','startdate','start_date'],
    'bill_end':             ['billend','bill_end','enddate','end_date'],
    'flag_created':         ['flagcreated','flag_created','flagdate'],
    'status':               ['status','flagstatus','flag_status'],
    'issue_type':           ['issue_type','issuetype','flagissue','flag_issue'],
    'flag_assignee':        ['flag_assignee','flagassignee','assignee'],
    'flag_category':        ['flag_category','flagcategory','category'],
    'pending_days':         ['pending_days','pendingdays','flag#days','days_open'],
    'bucket_days':          ['bucket_days','bucket_Days','bucketdays'],
    'responsibility':       ['issue_type_responsibility','Issue_Type_Responsibility','responsibility'],
    'assignee_org':         ['flag assigneeorg','Flag AssigneeOrg','assigneeorg'],
    'notes':                ['notes','Notes','note'],
    'batch_type':           ['batch_type','Batch_Type','batchtype'],
    'live_flag':            ['historical/live','Historical/Live','live_flag'],
    'batch_code':           ['batchcode','batch_code'],
    'vendor_name':          ['vendor','vendor_name','vendorname'],
    'cost':                 ['cost','totalcost','total_cost','bill_cost'],
}

def _resolve_columns(columns: list) -> dict:
    """Map actual column names to canonical names using alias table."""
    col_lower = {c.lower().strip(): c for c in columns}
    resolved  = {}
    for canonical, aliases in _COL_ALIASES.items():
        for alias in aliases:
            if alias.lower() in col_lower:
                resolved[canonical] = col_lower[alias.lower()]
                break
    return resolved

def _is_flat_flag_format(wb) -> bool:
    """Return True if workbook looks like the flat custom flag format."""
    try:
        ws = wb.worksheets[0]
        first_row = [str(v).strip().lower() if v is not None else ''
                     for v in next(ws.iter_rows(values_only=True))]
        # Must have billid/bill_id and issue_type/issuetype columns
        has_billid    = any(c in first_row for c in ['billid','bill_id','bill id'])
        has_issue     = any(c in first_row for c in ['issue_type','issuetype','flag_issue'])
        # R27 format has many sheets (one per bill); custom format has 1-2 sheets
        is_single_sheet = len(wb.sheetnames) <= 2
        return has_billid and has_issue and is_single_sheet
    except Exception:
        return False

def parse_custom_flags(path: str) -> pd.DataFrame:
    """
    Parse a flat/tabular custom flag export (one row per issue) into the same
    normalized DataFrame schema as parse_report27_text.

    Extra columns present in this format (not in R27) are preserved:
      pending_days, bucket_days, responsibility, assignee_org,
      notes, batch_type, flag_category, batch_code
    """
    wb  = load_workbook(path, data_only=True)
    ws  = wb.worksheets[0]
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return pd.DataFrame()

    # Parse header
    raw_header = [str(v).strip() if v is not None else '' for v in all_rows[0]]
    col_map    = _resolve_columns(raw_header)

    # Build raw DataFrame
    raw_df = pd.DataFrame(all_rows[1:], columns=raw_header)

    def col(canonical):
        """Get column values by canonical name, return None series if missing."""
        actual = col_map.get(canonical)
        if actual and actual in raw_df.columns:
            return raw_df[actual]
        return pd.Series([None] * len(raw_df))

    # ── Normalise to one-row-per-issue intermediate ────────────────────────
    interim = pd.DataFrame({
        'bill_id':       col('bill_id').astype(str).str.strip(),
        'account':       col('account').astype(str).str.strip(),
        'billing_period':col('billing_period').astype(str).str.strip(),
        'status':        col('status').fillna('Unresolved').astype(str).str.strip(),
        'issue_type':    col('issue_type').astype(str).str.strip(),
        'flag_created':  col('flag_created'),
        'vendor':        col('vendor_name').astype(str).str.strip()
                         if col_map.get('vendor_name') else pd.Series([''] * len(raw_df)),
        'cost':          pd.to_numeric(col('cost'), errors='coerce').fillna(0.0),
        'assigned_to':   col('flag_assignee').astype(str).str.strip(),
        'flag_category': col('flag_category').astype(str).str.strip(),
        'pending_days':  pd.to_numeric(col('pending_days'), errors='coerce'),
        'bucket_days':   col('bucket_days').astype(str).str.strip(),
        'responsibility':col('responsibility').astype(str).str.strip(),
        'assignee_org':  col('assignee_org').astype(str).str.strip(),
        'notes':         col('notes').astype(str).str.strip(),
        'batch_type':    col('batch_type').astype(str).str.strip(),
        'batch_code':    col('batch_code').astype(str).str.strip(),
    })

    # Clean issue_type trailing whitespace (known issue in source data)
    interim['issue_type'] = interim['issue_type'].str.strip()

    # ── Aggregate: one row per bill, issues collapsed into list ───────────
    # Aggregate: one row per bill using vectorised groupby
    # Issues are collected as sorted unique lists per bill
    issues_agg = (interim.groupby('bill_id')['issue_type']
                         .apply(lambda x: sorted(x.str.strip().unique().tolist()))
                         .rename('issues_list'))

    # First-row fields per bill (take first occurrence)
    first_cols = ['account','vendor','billing_period','status',
                  'cost','assigned_to','flag_created','bucket_days',
                  'responsibility','assignee_org','notes','batch_type','batch_code']
    first_cols = [c for c in first_cols if c in interim.columns]
    first_df = interim.groupby('bill_id')[first_cols].first()

    # Max pending_days per bill
    pending_max = interim.groupby('bill_id')['pending_days'].max()

    # flag_category: join unique values per bill
    cat_agg = (interim.groupby('bill_id')['flag_category']
                      .apply(lambda x: ', '.join(sorted(x.unique().tolist())))
                      .rename('flag_category'))

    # Combine
    df = (first_df
          .join(issues_agg)
          .join(pending_max)
          .join(cat_agg)
          .reset_index())

    # Derived fields from issues_list
    df['flag_issues']     = df['issues_list'].apply(lambda x: ', '.join(x))
    df['num_issues']      = df['issues_list'].apply(len)
    df['flagged_date']    = pd.to_datetime(df.get('flag_created'), errors='coerce')
    df['resolved_date']   = None
    df['resolvers']       = [[] for _ in range(len(df))]
    df['days_to_resolve'] = None
    df['primary_resolver']= 'Unresolved'
    df['cost_recovery']   = 0.0
    df['address']         = ''
    df['flag_events']     = [[] for _ in range(len(df))]

    # ── Derived columns (same as R27 output) ─────────────────────────────
    df['billing_period_dt']    = pd.to_datetime(
        df['billing_period'].str.strip(), format='%Y%m', errors='coerce')
    df['billing_period_label'] = df['billing_period_dt'].dt.strftime('%b %Y')
    df['primary_issue']        = df['issues_list'].apply(lambda x: x[0] if x else 'Unknown')
    df['primary_category']     = df['primary_issue'].apply(
        lambda x: FLAG_META.get(x, {}).get('category', 'Other'))
    df['primary_priority']     = df['primary_issue'].apply(
        lambda x: FLAG_META.get(x, {}).get('priority', 'Medium'))

    # Flag source marker — lets app show format-specific fields
    df['flag_source'] = 'custom'

    return df


def detect_and_parse_flags(path: str) -> tuple:
    """
    Auto-detect the flag file format and parse accordingly.
    Returns (DataFrame, format_name) where format_name is 'r27' or 'custom'.
    """
    wb = load_workbook(path, data_only=True)

    if _is_flat_flag_format(wb):
        return parse_custom_flags(path), 'custom'
    else:
        # R27 format — use existing text-based parser via openpyxl fallback
        lines = []
        for ws in wb.worksheets:
            if ws.title.lower() in {'report overview', 'overview'}:
                continue
            for row in ws.iter_rows(values_only=True):
                lines.append('\t'.join(str(c) if c is not None else '' for c in row))
        text = '\n'.join(lines)
        return parse_report27_text(text), 'r27'
