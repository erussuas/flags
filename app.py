import io
from collections import Counter

import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st

from ec_parser import (
    FLAG_META, PRIORITY_ORDER, PRIORITY_COLOR, CATEGORY_COLOR,
    GHG_COMMODITIES, GHG_SCOPE, DISPLAY_UNITS, FROM_KWH,
    from_kwh, kwh,
    load_report27, load_report18, load_report03,
)

# ══════════════════════════════════════════════════════════════════════════════
# PAGE CONFIG
# ══════════════════════════════════════════════════════════════════════════════
st.set_page_config(
    page_title="EnergyCAP Bill Flag Analyzer",
    page_icon="⚑", layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
  .metric-card{background:#f8f9fa;border-radius:10px;padding:1rem 1.25rem;border:1px solid #e9ecef;}
  .metric-label{font-size:12px;color:#6c757d;margin:0 0 4px 0;}
  .metric-value{font-size:26px;font-weight:600;color:#212529;margin:0;}
  .metric-sub{font-size:12px;color:#6c757d;margin:4px 0 0 0;}
  .action-card{border-left:4px solid;border-radius:6px;padding:.75rem 1rem;margin-bottom:.6rem;background:#fafafa;}
  .action-high{border-color:#dc3545;background:#fff5f5;}
  .action-medium{border-color:#fd7e14;background:#fff8f0;}
  .action-low{border-color:#198754;background:#f0fff4;}
  .action-info{border-color:#0d6efd;background:#f0f6ff;}
  .bill-card{background:#fff;border:1px solid #dee2e6;border-radius:8px;padding:.85rem 1rem;margin-bottom:.5rem;}
  .bill-card-header{display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:6px;}
  .bill-id{font-weight:600;font-size:14px;color:#212529;}
  .bill-cost{font-weight:600;font-size:15px;color:#0d6efd;}
  .bill-meta{font-size:12px;color:#6c757d;margin:2px 0;}
  .flag-badge{display:inline-block;font-size:11px;padding:2px 8px;border-radius:12px;font-weight:500;margin:2px;}
  .badge-red{background:#ffe0e0;color:#c0392b;} .badge-orange{background:#fff0d6;color:#9b5504;}
  .badge-green{background:#e0f7ea;color:#1a7a4a;} .badge-blue{background:#ddeeff;color:#1a4fa0;}
  .badge-gray{background:#e9ecef;color:#495057;} .badge-purple{background:#f0e8ff;color:#5a1fa0;}
  .drill-banner{background:#e8f0fe;border:1px solid #4285f4;border-radius:8px;
    padding:.6rem 1rem;margin-bottom:1rem;display:flex;align-items:center;justify-content:space-between;}
  .drill-label{font-size:13px;font-weight:500;color:#1a56db;}
  .upload-section{background:#f8f9fa;border-radius:8px;padding:.75rem 1rem;margin-bottom:.5rem;
    border:1px solid #e9ecef;}
  .upload-ok{color:#198754;font-size:12px;font-weight:500;}
  .upload-warn{color:#fd7e14;font-size:12px;font-weight:500;}
</style>
""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# CACHING — all heavy work done once, keyed by file bytes
# ══════════════════════════════════════════════════════════════════════════════
@st.cache_data(show_spinner=False)
def cached_load_report03(file_bytes: bytes) -> pd.DataFrame:
    import tempfile, os
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as f:
        f.write(file_bytes); path = f.name
    try:
        from ec_parser import parse_report03
        return parse_report03(path)
    finally:
        os.unlink(path)

@st.cache_data(show_spinner=False)
def cached_load_report18(file_bytes: bytes, setup_bytes: bytes | None) -> pd.DataFrame:
    import tempfile, os
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as f:
        f.write(file_bytes); path18 = f.name
    setup_df = None
    if setup_bytes:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as f:
            f.write(setup_bytes); path03 = f.name
        try:
            from ec_parser import parse_report03
            setup_df = parse_report03(path03)
        finally:
            os.unlink(path03)
    try:
        from ec_parser import parse_report18
        return parse_report18(path18, setup_df)
    finally:
        os.unlink(path18)

@st.cache_data(show_spinner=False)
def cached_load_report27(file_bytes: bytes) -> pd.DataFrame:
    import tempfile, os, shutil, subprocess
    from ec_parser import parse_report27_text
    from openpyxl import load_workbook
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as f:
        f.write(file_bytes); path = f.name
    try:
        text = ""
        if shutil.which("extract-text"):
            r = subprocess.run(["extract-text", path], capture_output=True, text=True, timeout=60)
            text = r.stdout
        if not text.strip():
            wb = load_workbook(path, data_only=True)
            lines = []
            for ws in wb.worksheets:
                # Skip metadata/overview sheets
                if ws.title.lower() in {"report overview", "overview"}:
                    continue
                for row in ws.iter_rows(values_only=True):
                    # Use tab separator so "Account:		Name" is preserved correctly
                    cells = [str(c) if c is not None else "" for c in row]
                    lines.append("\t".join(cells))
            text = "\n".join(lines)
        return parse_report27_text(text)
    finally:
        os.unlink(path)

@st.cache_data(show_spinner=False)
def build_master(r27_bytes_list: list[bytes], r18_bytes_list: list[bytes],
                 r03_bytes: bytes | None) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Build merged DataFrames. Called only when uploaded files change."""
    # Parse R03 once
    setup_df = cached_load_report03(r03_bytes) if r03_bytes else pd.DataFrame()
    setup_bytes = r03_bytes  # pass bytes, not df, to cached R18 loader

    # Merge all R27 files
    r27_frames = [cached_load_report27(b) for b in r27_bytes_list]
    df27 = pd.concat(r27_frames, ignore_index=True).drop_duplicates("bill_id") if r27_frames else pd.DataFrame()

    # Merge all R18 files
    r18_frames = [cached_load_report18(b, setup_bytes) for b in r18_bytes_list]
    df18 = pd.concat(r18_frames, ignore_index=True).drop_duplicates("bill_id") if r18_frames else pd.DataFrame()

    # Join R27 + R18 on bill_id
    if not df27.empty and not df18.empty:
        df_merged = df27.merge(df18, on="bill_id", how="left")
    else:
        df_merged = df27.copy() if not df27.empty else pd.DataFrame()

    return df_merged, setup_df, df18


# ══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════════════════
PT = dict(font_family="Inter,system-ui,sans-serif",
          paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)")

SORT_OPTS = ["Cost (high→low)","Cost (low→high)",
             "Usage kWh (high→low)","Days to resolve (longest)","Bill ID","Vendor A→Z"]
SORT_MAP  = {
    "Cost (high→low)":          ("cost",False),
    "Cost (low→high)":          ("cost",True),
    "Usage kWh (high→low)":     ("total_kwh_equivalent",False),
    "Days to resolve (longest)":("days_to_resolve",False),
    "Bill ID":                  ("bill_id",True),
    "Vendor A→Z":               ("vendor",True),
}

def metric_html(label, value, sub="", color="#212529"):
    return f"""<div class="metric-card"><p class="metric-label">{label}</p>
    <p class="metric-value" style="color:{color}">{value}</p>
    <p class="metric-sub">{sub}</p></div>"""

def badge(text, color="gray"):
    return f'<span class="flag-badge badge-{color}">{text}</span>'

def priority_icon(p):
    return {"High":"🔴","Medium":"🟠","Low":"🟢"}.get(p,"⚪")

def bar_h(data, title, color="#0d6efd", height=320, highlight=None):
    labels, values = list(data.keys()), list(data.values())
    colors = ["#ff6b35" if highlight and l==highlight else color for l in labels]
    fig = go.Figure(go.Bar(x=values,y=labels,orientation="h",marker_color=colors,
                           text=values,textposition="outside",
                           hovertemplate="%{y}: %{x}<extra></extra>"))
    fig.update_layout(title=title,height=height,yaxis_autorange="reversed",
                      xaxis=dict(showgrid=True,gridcolor="#f0f0f0"),
                      margin=dict(l=10,r=50,t=40,b=10),
                      showlegend=False,clickmode="event+select",**PT)
    return fig

def bar_v(data, title, color="#0d6efd", height=320, highlight=None):
    labels, values = list(data.keys()), list(data.values())
    colors = ["#ff6b35" if highlight and l==highlight else color for l in labels]
    fig = go.Figure(go.Bar(x=labels,y=values,marker_color=colors,
                           text=values,textposition="outside",
                           hovertemplate="%{x}: %{y}<extra></extra>"))
    fig.update_layout(title=title,height=height,
                      yaxis=dict(showgrid=True,gridcolor="#f0f0f0"),
                      margin=dict(l=10,r=10,t=40,b=10),
                      showlegend=False,clickmode="event+select",**PT)
    return fig

def donut(data, title, colors=None, height=280):
    colors = colors or px.colors.qualitative.Set2
    fig = go.Figure(go.Pie(labels=list(data.keys()),values=list(data.values()),
                           hole=0.55,marker_colors=colors,
                           textinfo="percent+label",textfont_size=11,
                           hovertemplate="%{label}: %{value}<extra></extra>"))
    fig.update_layout(title=title,height=height,
                      margin=dict(l=10,r=10,t=40,b=10),showlegend=False,**PT)
    return fig

def extract_click(ev):
    if not ev: return None
    pts = ev.get("selection",{}).get("points",[])
    if not pts: return None
    return pts[0].get("y") or pts[0].get("x") or None

def render_bill_cards(sub_df, display_unit="kWh", max_cards=50):
    has_usage = "total_kwh_equivalent" in sub_df.columns
    if sub_df.empty:
        st.info("No bills match this selection.")
        return
    st.caption(f"Showing {min(len(sub_df),max_cards)} of {len(sub_df)} bills")
    for _, row in sub_df.head(max_cards).iterrows():
        p  = row.get("primary_priority","Medium")
        bc = {"High":"red","Medium":"orange","Low":"green"}.get(p,"blue")
        sc = "#198754" if row["status"]=="Resolved" else "#dc3545"
        days = f"{int(row['days_to_resolve'])}d" if pd.notna(row.get("days_to_resolve")) else "—"
        resolver = (row["resolvers"][-1] if row["resolvers"] else "—")
        usage_str = ""
        if has_usage and pd.notna(row.get("total_kwh_equivalent")):
            val = from_kwh(row["total_kwh_equivalent"], display_unit)
            cat = row.get("ghg_category","")
            scope = row.get("ghg_scope","")
            usage_str = (f"&nbsp;·&nbsp; <strong>{val:,.1f} {display_unit}</strong>"
                         f"{' · '+cat if cat else ''}{' ('+scope+')' if scope else ''}")
        badges = " ".join(badge(i, bc) for i in row["issues_list"])
        st.markdown(f"""<div class="bill-card">
          <div class="bill-card-header">
            <span class="bill-id">Bill {row['bill_id']} &nbsp;·&nbsp; {row['vendor']}</span>
            <span class="bill-cost">${row['cost']:,.2f}</span>
          </div>
          <div class="bill-meta">
            {row.get('account','')}<br>
            Period: {row.get('billing_period_label','—')} &nbsp;·&nbsp;
            Assigned: {row.get('assigned_to','—')} &nbsp;·&nbsp;
            Resolved by: {resolver} &nbsp;·&nbsp;
            Days: {days} &nbsp;·&nbsp;
            <span style="color:{sc};font-weight:500">{row['status']}</span>
            {usage_str}
          </div>
          <div style="margin-top:5px">{badges}</div>
        </div>""", unsafe_allow_html=True)

def drill_banner(label, key):
    c1, c2 = st.columns([11,1])
    with c1:
        st.markdown(f'<div class="drill-banner"><span class="drill-label">'
                    f'🔍 Drill-down: <strong>{label}</strong></span></div>',
                    unsafe_allow_html=True)
    with c2:
        if st.button("✕", key=f"clr_{key}"):
            st.session_state[key] = None
            st.rerun()

def bill_detail_panel(row, flag_meta=FLAG_META):
    st.subheader(f"Bill {row['bill_id']} — {row['vendor']}")
    c1, c2 = st.columns(2)
    with c1:
        st.markdown(f"**Account:** {row['account']}")
        st.markdown(f"**Period:** {row.get('billing_period_label','—')}")
        st.markdown(f"**Cost:** ${row['cost']:,.2f}  |  **Recovery:** ${row['cost_recovery']:,.2f}")
        st.markdown(f"**Assigned to:** {row.get('assigned_to','—')}")
        if pd.notna(row.get("total_kwh_equivalent")):
            st.markdown(f"**Total GHG usage:** {row['total_kwh_equivalent']:,.0f} kWh equivalent")
            if row.get("ghg_category"):
                st.markdown(f"**Commodity:** {row['ghg_category']} ({row.get('ghg_scope','')})")
    with c2:
        st.markdown(f"**Status:** {row['status']}")
        days = f"{int(row['days_to_resolve'])}d" if pd.notna(row.get('days_to_resolve')) else "Pending"
        st.markdown(f"**Days to resolve:** {days}")
        if row.get("resolvers"):
            st.markdown(f"**Resolved by:** {', '.join(row['resolvers'])}")
    st.markdown("**Flag issues & guidance:**")
    for issue in row["issues_list"]:
        m = flag_meta.get(issue, {})
        icon = priority_icon(m.get("priority",""))
        with st.expander(f"{icon} {issue} — {m.get('category','')}"):
            if m:
                st.markdown(f"**Cause:** {m['cause']}")
                st.markdown(f"**Action:** {m['action']}")
                st.markdown(f"**In EnergyCAP:** {m['in_energycap']}")


# ══════════════════════════════════════════════════════════════════════════════
# SESSION STATE INIT
# ══════════════════════════════════════════════════════════════════════════════
def ss_init():
    defaults = {
        "drill_vendor":None,"drill_issue":None,"drill_period":None,
        "drill_assignee":None,"active_tab":"overview",
        "ft_issues":[],"ft_vendors":[],"ft_assignees":[],
        "ft_status":["Resolved","Unresolved"],"ft_priority":["High","Medium","Low"],
        "ft_ghg_cat":[],"ft_sort":"Cost (high→low)",
    }
    for k,v in defaults.items():
        if k not in st.session_state: st.session_state[k] = v
ss_init()


# ══════════════════════════════════════════════════════════════════════════════
# SIDEBAR — UPLOAD + GLOBAL FILTERS
# ══════════════════════════════════════════════════════════════════════════════
with st.sidebar:
    st.title("⚑ Bill Flag Analyzer")
    st.caption("EnergyCAP Report-27 · Report-18 · Report-03")
    st.divider()

    # ── Upload section ────────────────────────────────────────────────────────
    st.markdown("### Upload Files")
    st.caption("Upload any combination. Multiple Report-27 and Report-18 files are merged automatically.")

    with st.expander("📂 Report-27 — Bill Flags (required)", expanded=True):
        r27_files = st.file_uploader("One or more Report-27 exports",
                                     type=["xlsx"], accept_multiple_files=True,
                                     key="r27_upload",
                                     help="Bills → Menu (≡) → Report-27 Bill Flags → Export to Excel")
        if r27_files:
            for f in r27_files:
                st.markdown(f'<span class="upload-ok">✓ {f.name}</span>', unsafe_allow_html=True)

    with st.expander("📂 Report-18 — Bill Line Items (optional)", expanded=True):
        r18_files = st.file_uploader("One or more Report-18 exports (Use sheet)",
                                     type=["xlsx"], accept_multiple_files=True,
                                     key="r18_upload",
                                     help="Bills → Report-18 Bill Line Item Report, filter Line Type = Use")
        if r18_files:
            for f in r18_files:
                st.markdown(f'<span class="upload-ok">✓ {f.name}</span>', unsafe_allow_html=True)
        else:
            st.markdown('<span class="upload-warn">ⓘ Without Report-18, GHG usage data unavailable</span>',
                        unsafe_allow_html=True)

    with st.expander("📂 Report-03 — Setup / Master Data (recommended)", expanded=True):
        r03_file = st.file_uploader("Report-03 Setup Report",
                                    type=["xlsx"], accept_multiple_files=False,
                                    key="r03_upload",
                                    help="All Reports → Setup Report for Accounts, Vendors, Cost Centers, Meters, Sites (Excel only)")
        if r03_file:
            st.markdown(f'<span class="upload-ok">✓ {r03_file.name}</span>', unsafe_allow_html=True)
        else:
            st.markdown('<span class="upload-warn">ⓘ Without Report-03, commodity classification uses caption heuristics</span>',
                        unsafe_allow_html=True)

    st.divider()
    st.markdown("### Global Filters")

    # Placeholder filters — populated after data loads
    if "df_master" in st.session_state and not st.session_state.df_master.empty:
        df_all = st.session_state.df_master
        all_vendors = sorted(df_all["vendor"].unique())
        status_filter   = st.multiselect("Status", ["Resolved","Unresolved"],
                                         default=["Resolved","Unresolved"])
        vendor_filter   = st.multiselect("Vendor", all_vendors, default=all_vendors)
        priority_filter = st.multiselect("Priority",["High","Medium","Low"],
                                         default=["High","Medium","Low"])
    else:
        status_filter   = ["Resolved","Unresolved"]
        vendor_filter   = []
        priority_filter = ["High","Medium","Low"]

    st.divider()
    st.caption("Click any bar in a chart to drill into matching bills. "
               "Click ✕ to clear the drill-down.")


# ══════════════════════════════════════════════════════════════════════════════
# LOAD & MERGE DATA (cached)
# ══════════════════════════════════════════════════════════════════════════════
if not r27_files:
    st.title("EnergyCAP Bill Flag Analyzer")
    st.info(
        "👈 Upload at least one **Report-27 Bill Flags** export to get started.\n\n"
        "**Optional but recommended:**\n"
        "- **Report-18** (Bill Line Item Report, Use sheet) — adds GHG usage data\n"
        "- **Report-03** (Setup Report) — enables authoritative commodity classification\n\n"
        "Multiple files of the same type are merged automatically."
    )
    st.stop()

# Read bytes (Streamlit UploadedFile is only readable once)
r27_bytes_list = [f.read() for f in r27_files]
r18_bytes_list = [f.read() for f in r18_files] if r18_files else []
r03_bytes      = r03_file.read() if r03_file else None

with st.spinner("Parsing and merging files… (this only happens when files change)"):
    df_master, df_setup, df18_full = build_master(r27_bytes_list, r18_bytes_list, r03_bytes)

st.session_state.df_master = df_master

if df_master.empty:
    st.warning("No bill records found in the uploaded files.")
    st.stop()

has_usage  = "total_kwh_equivalent" in df_master.columns and df_master["total_kwh_equivalent"].notna().any()
has_setup  = not df_setup.empty
all_vendors = sorted(df_master["vendor"].unique())
if not vendor_filter: vendor_filter = all_vendors

# Apply global filters
df = df_master.copy()
if status_filter:   df = df[df["status"].isin(status_filter)]
if vendor_filter:   df = df[df["vendor"].isin(vendor_filter)]
if priority_filter: df = df[df["primary_priority"].isin(priority_filter)]

if df.empty:
    st.warning("No records match the current filters.")
    st.stop()

# Derived
issues_exp = (df.explode("issues_list").rename(columns={"issues_list":"issue"})
               .query("issue.notna() and issue != ''"))
issue_counts  = Counter(issues_exp["issue"].tolist())
vendor_counts = df["vendor"].value_counts().head(12).to_dict()
period_counts = (df[df["billing_period_label"].notna()]
                 .groupby("billing_period_label").size().sort_index().to_dict())
assignee_rows = [a.strip()
                 for _,r in df.iterrows()
                 for a in str(r["assigned_to"]).split(",") if a.strip()]
resolver_rows = [r for _,row in df.iterrows() for r in row["resolvers"]]
all_assignees = sorted(set(assignee_rows))
all_issues    = sorted(issue_counts.keys())

total_bills      = len(df)
resolved_ct      = (df["status"]=="Resolved").sum()
unresolved_ct    = (df["status"]=="Unresolved").sum()
resolution_rate  = resolved_ct/total_bills*100 if total_bills else 0
total_cost       = df["cost"].sum()
total_recovery   = df["cost_recovery"].sum()
multi_issue_ct   = (df["num_issues"]>=3).sum()
avg_resolve      = df["days_to_resolve"].dropna().mean()
hp_open          = df[(df["primary_priority"]=="High")&(df["status"]=="Unresolved")].shape[0]
total_kwh        = df["total_kwh_equivalent"].sum() if has_usage else 0
ghg_bills_ct     = df["total_kwh_equivalent"].notna().sum() if has_usage else 0


# ══════════════════════════════════════════════════════════════════════════════
# TABS
# ══════════════════════════════════════════════════════════════════════════════
t_overview, t_flags, t_vendors, t_ghg, t_actions, t_detail = st.tabs([
    "📊 Overview", "⚑ Flag Analysis", "🏢 Vendors",
    "🌿 GHG Priority Queue", "✅ Action Guide", "📋 Bill Detail",
])


# ══════════════════════════════════════════════════════════════════════════════
# TAB 1 — OVERVIEW
# ══════════════════════════════════════════════════════════════════════════════
with t_overview:
    st.subheader("Summary")
    cols = st.columns(6)
    metrics = [
        ("Total flagged bills",    str(total_bills),
         f"{resolved_ct} resolved · {unresolved_ct} open","#212529"),
        ("Resolution rate",        f"{resolution_rate:.0f}%",
         f"{resolved_ct} of {total_bills}",
         "#198754" if resolution_rate>=90 else "#fd7e14"),
        ("Total bill value",       f"${total_cost:,.0f}","Under review","#212529"),
        ("Cost recovered",         f"${total_recovery:,.0f}","Tracked savings","#198754"),
        ("GHG bills w/ usage",
         str(ghg_bills_ct) if has_usage else "—",
         f"{total_kwh/1e6:,.1f}M kWh equiv." if has_usage else "Upload Report-18","#6f42c1"),
        ("High-priority open",     str(hp_open),"Need immediate action",
         "#dc3545" if hp_open>0 else "#198754"),
    ]
    for col,(label,val,sub,color) in zip(cols,metrics):
        col.markdown(metric_html(label,val,sub,color),unsafe_allow_html=True)
    st.markdown("")

    # Flag issues chart (clickable)
    col1,col2 = st.columns([3,2])
    with col1:
        top12 = dict(sorted(issue_counts.items(),key=lambda x:-x[1])[:12])
        colors_i = [PRIORITY_COLOR.get(FLAG_META.get(k,{}).get("priority","Medium"),"#6c757d")
                    for k in top12]
        fig = go.Figure(go.Bar(x=list(top12.values()),y=list(top12.keys()),
                               orientation="h",marker_color=colors_i,
                               text=list(top12.values()),textposition="outside",
                               hovertemplate="%{y}: %{x}<extra></extra>"))
        fig.update_layout(title="Flag issues — click to drill",yaxis_autorange="reversed",
                          xaxis=dict(showgrid=True,gridcolor="#f0f0f0"),
                          height=400,margin=dict(l=10,r=50,t=40,b=10),
                          showlegend=False,clickmode="event+select",**PT)
        ev1 = st.plotly_chart(fig,use_container_width=True,on_select="rerun",key="ov_issues")
        ci  = extract_click(ev1)
        if ci and ci in issue_counts:
            st.session_state.drill_issue = ci
            st.session_state.active_tab  = "overview"

    with col2:
        if period_counts:
            ev2 = st.plotly_chart(
                bar_v(period_counts,"Bills by billing period — click to drill",
                      color="#0d6efd",height=400,
                      highlight=st.session_state.drill_period),
                use_container_width=True,on_select="rerun",key="ov_period")
            cp = extract_click(ev2)
            if cp and cp in period_counts:
                st.session_state.drill_period = cp
                st.session_state.active_tab   = "overview"

    # Drill panels
    if st.session_state.drill_issue and st.session_state.active_tab=="overview":
        di = st.session_state.drill_issue
        drill_banner(f"Issue: {di}","drill_issue")
        sub = df[df["issues_list"].apply(lambda lst: di in lst)]
        m = FLAG_META.get(di,{})
        if m:
            with st.expander(f"📖 About: {di}",expanded=False):
                st.markdown(f"**Cause:** {m['cause']}")
                st.markdown(f"**Action:** {m['action']}")
        render_bill_cards(sub.sort_values("cost",ascending=False))

    if st.session_state.drill_period and st.session_state.active_tab=="overview":
        dp = st.session_state.drill_period
        drill_banner(f"Period: {dp}","drill_period")
        render_bill_cards(df[df["billing_period_label"]==dp].sort_values("cost",ascending=False))

    # Donuts
    c3,c4,c5 = st.columns(3)
    with c3:
        cat_ct = (issues_exp["issue"]
                  .apply(lambda x: FLAG_META.get(x,{}).get("category","Other"))
                  .value_counts().to_dict())
        st.plotly_chart(donut(cat_ct,"Issues by category",
                              [CATEGORY_COLOR.get(k,"#888") for k in cat_ct]),
                        use_container_width=True)
    with c4:
        st.plotly_chart(donut(df["status"].value_counts().to_dict(),
                              "Flag status",["#198754","#dc3545"]),
                        use_container_width=True)
    with c5:
        pri_ct = (issues_exp["issue"]
                  .apply(lambda x: FLAG_META.get(x,{}).get("priority","Medium"))
                  .value_counts().reindex(["High","Medium","Low"]).dropna().to_dict())
        st.plotly_chart(donut(pri_ct,"Issues by priority",
                              ["#dc3545","#fd7e14","#198754"]),
                        use_container_width=True)

    # Resolution time
    res_df = df[df["days_to_resolve"].notna()&(df["days_to_resolve"]>=0)]
    if not res_df.empty:
        st.subheader("Resolution time distribution")
        fig_r = px.histogram(res_df,x="days_to_resolve",nbins=20,
                             labels={"days_to_resolve":"Days to resolve"},
                             color_discrete_sequence=["#0d6efd"],
                             title=f"Avg {avg_resolve:.1f} d · Median {res_df['days_to_resolve'].median():.0f} d")
        fig_r.update_layout(height=240,margin=dict(l=10,r=10,t=40,b=10),showlegend=False,**PT)
        st.plotly_chart(fig_r,use_container_width=True)


# ══════════════════════════════════════════════════════════════════════════════
# TAB 2 — FLAG ANALYSIS
# ══════════════════════════════════════════════════════════════════════════════
with t_flags:
    st.subheader("Flag Analysis")

    # Multi-filter panel
    with st.expander("🔽 Filter flags",expanded=True):
        fc1,fc2,fc3 = st.columns(3)
        with fc1:
            ft_issues = st.multiselect("Flag issue type",all_issues,
                                       default=st.session_state.ft_issues,
                                       placeholder="All issues",key="ft_issues_w")
            ft_status = st.multiselect("Status",["Resolved","Unresolved"],
                                       default=st.session_state.ft_status,key="ft_status_w")
        with fc2:
            ft_vendors = st.multiselect("Vendor",all_vendors,
                                        default=st.session_state.ft_vendors,
                                        placeholder="All vendors",key="ft_vendors_w")
            ft_priority = st.multiselect("Priority",["High","Medium","Low"],
                                         default=st.session_state.ft_priority,key="ft_pri_w")
        with fc3:
            ft_assignees = st.multiselect("Assignee",all_assignees,
                                          default=st.session_state.ft_assignees,
                                          placeholder="All assignees",key="ft_assign_w")
            if has_usage:
                ghg_cats = sorted(df["ghg_category"].dropna().unique().tolist())
                ghg_cats = [g for g in ghg_cats if g]
                ft_ghg = st.multiselect("GHG category",ghg_cats,
                                        default=st.session_state.ft_ghg_cat,
                                        placeholder="All commodities",key="ft_ghg_w")
            else:
                ft_ghg = []
            ft_sort = st.selectbox("Sort bills by",SORT_OPTS,
                                   index=SORT_OPTS.index(st.session_state.ft_sort),
                                   key="ft_sort_w")

        # Persist
        st.session_state.ft_issues    = ft_issues
        st.session_state.ft_vendors   = ft_vendors
        st.session_state.ft_assignees = ft_assignees
        st.session_state.ft_status    = ft_status
        st.session_state.ft_priority  = ft_priority
        st.session_state.ft_ghg_cat   = ft_ghg
        st.session_state.ft_sort      = ft_sort

        _,rb = st.columns([8,2])
        with rb:
            if st.button("↺ Reset all filters",use_container_width=True):
                for k in ["ft_issues","ft_vendors","ft_assignees","ft_ghg_cat"]:
                    st.session_state[k] = []
                st.session_state.ft_status   = ["Resolved","Unresolved"]
                st.session_state.ft_priority = ["High","Medium","Low"]
                st.session_state.ft_sort     = "Cost (high→low)"
                st.rerun()

    # Apply in-tab filters
    fdf = df.copy()
    if ft_issues:
        fdf = fdf[fdf["issues_list"].apply(lambda lst: any(i in lst for i in ft_issues))]
    if ft_vendors:   fdf = fdf[fdf["vendor"].isin(ft_vendors)]
    if ft_status:    fdf = fdf[fdf["status"].isin(ft_status)]
    if ft_priority:  fdf = fdf[fdf["primary_priority"].isin(ft_priority)]
    if ft_assignees:
        fdf = fdf[fdf["assigned_to"].apply(
            lambda a: any(x.strip() in str(a) for x in ft_assignees))]
    if ft_ghg and has_usage:
        fdf = fdf[fdf["ghg_category"].isin(ft_ghg)]

    sc_f,sa_f = SORT_MAP[ft_sort]
    fdf = fdf.sort_values(sc_f,ascending=sa_f,na_position="last")

    fi_exp = (fdf.explode("issues_list").rename(columns={"issues_list":"issue"})
               .query("issue.notna() and issue != ''"))

    # Summary metrics
    fm1,fm2,fm3,fm4 = st.columns(4)
    fm1.metric("Bills shown",  len(fdf))
    fm2.metric("Total cost",   f"${fdf['cost'].sum():,.0f}")
    fm3.metric("Unresolved",   (fdf["status"]=="Unresolved").sum())
    fm4.metric("Unique issues",fi_exp["issue"].nunique())

    # Charts (clickable)
    ch1,ch2 = st.columns(2)
    with ch1:
        fi_ct = fi_exp["issue"].value_counts().head(12).to_dict()
        if fi_ct:
            ev_fi = st.plotly_chart(
                bar_h(fi_ct,"Issues — click to drill",color="#0d6efd",height=320,
                      highlight=st.session_state.drill_issue),
                use_container_width=True,on_select="rerun",key="ft_issues_chart")
            cfi = extract_click(ev_fi)
            if cfi and cfi in fi_ct:
                st.session_state.drill_issue = cfi
                st.session_state.active_tab  = "flags"

    with ch2:
        fv_ct = fdf["vendor"].value_counts().head(12).to_dict()
        if fv_ct:
            ev_fv = st.plotly_chart(
                bar_h(fv_ct,"Vendors — click to drill",color="#6f42c1",height=320,
                      highlight=st.session_state.drill_vendor),
                use_container_width=True,on_select="rerun",key="ft_vendors_chart")
            cfv = extract_click(ev_fv)
            if cfv and cfv in fv_ct:
                st.session_state.drill_vendor = cfv
                st.session_state.active_tab   = "flags"

    ch3,ch4 = st.columns(2)
    with ch3:
        vis_assignees = [a.strip() for _,r in fdf.iterrows()
                         for a in str(r["assigned_to"]).split(",") if a.strip()]
        fa_ct = dict(Counter(vis_assignees).most_common(10))
        if fa_ct:
            ev_fa = st.plotly_chart(
                bar_h(fa_ct,"Assignees — click to drill",color="#198754",height=300,
                      highlight=st.session_state.drill_assignee),
                use_container_width=True,on_select="rerun",key="ft_assign_chart")
            cfa = extract_click(ev_fa)
            if cfa and cfa in fa_ct:
                st.session_state.drill_assignee = cfa
                st.session_state.active_tab     = "flags"
    with ch4:
        fp_ct = (fi_exp["issue"]
                 .apply(lambda x: FLAG_META.get(x,{}).get("priority","Medium"))
                 .value_counts().reindex(["High","Medium","Low"]).dropna().to_dict())
        if fp_ct:
            st.plotly_chart(donut(fp_ct,"Priority breakdown",
                                  ["#dc3545","#fd7e14","#198754"],height=300),
                            use_container_width=True)

    # Drill panels
    if st.session_state.drill_issue and st.session_state.active_tab=="flags":
        di = st.session_state.drill_issue
        drill_banner(f"Issue: {di}","drill_issue")
        sub = fdf[fdf["issues_list"].apply(lambda lst: di in lst)]
        m = FLAG_META.get(di,{})
        if m:
            with st.expander(f"📖 About: {di}",expanded=False):
                st.markdown(f"**Cause:** {m['cause']}")
                st.markdown(f"**Action:** {m['action']}")
                st.markdown(f"**In EnergyCAP:** {m['in_energycap']}")
        render_bill_cards(sub)

    if st.session_state.drill_vendor and st.session_state.active_tab=="flags":
        dv = st.session_state.drill_vendor
        drill_banner(f"Vendor: {dv}","drill_vendor")
        render_bill_cards(fdf[fdf["vendor"]==dv])

    if st.session_state.drill_assignee and st.session_state.active_tab=="flags":
        da = st.session_state.drill_assignee
        drill_banner(f"Assignee: {da}","drill_assignee")
        render_bill_cards(fdf[fdf["assigned_to"].str.contains(da,na=False)])

    # Full list + row click
    st.divider()
    st.subheader(f"All matching bills ({len(fdf)})")
    disp_cols = {"bill_id":"Bill ID","vendor":"Vendor","billing_period_label":"Period",
                 "cost":"Cost ($)","status":"Status","flag_issues":"Flag Issues",
                 "assigned_to":"Assigned To","days_to_resolve":"Days",
                 "ghg_category":"Commodity","total_kwh_equivalent":"kWh Equiv."}
    show_cols = [c for c in disp_cols if c in fdf.columns]
    disp = fdf[show_cols].copy().rename(columns=disp_cols)
    disp["Cost ($)"] = disp["Cost ($)"].map("${:,.2f}".format)
    if "kWh Equiv." in disp.columns:
        disp["kWh Equiv."] = disp["kWh Equiv."].apply(
            lambda x: f"{x:,.0f}" if pd.notna(x) else "—")
    disp["Days"] = disp["Days"].apply(lambda x: f"{int(x)}d" if pd.notna(x) else "—")

    tbl = st.dataframe(disp,use_container_width=True,hide_index=True,height=380,
                       on_select="rerun",selection_mode="single-row",key="flags_table")
    sel = tbl.get("selection",{}).get("rows",[]) if tbl else []
    if sel:
        st.markdown("---")
        bill_detail_panel(fdf.iloc[sel[0]])

    buf = io.StringIO(); disp.to_csv(buf,index=False)
    st.download_button("⬇ Download as CSV",buf.getvalue(),"bill_flags_filtered.csv","text/csv")


# ══════════════════════════════════════════════════════════════════════════════
# TAB 3 — VENDORS
# ══════════════════════════════════════════════════════════════════════════════
with t_vendors:
    st.subheader("Vendor Analysis")
    col1,col2 = st.columns(2)
    with col1:
        ev_vc = st.plotly_chart(
            bar_h(vendor_counts,"Flags by vendor — click to drill",
                  color="#6f42c1",height=380,highlight=st.session_state.drill_vendor),
            use_container_width=True,on_select="rerun",key="vend_count")
        cv1 = extract_click(ev_vc)
        if cv1 and cv1 in vendor_counts:
            st.session_state.drill_vendor = cv1
            st.session_state.active_tab   = "vendors"
    with col2:
        vcost = df.groupby("vendor")["cost"].sum().sort_values(ascending=False).head(12).to_dict()
        ev_vco = st.plotly_chart(
            bar_h(vcost,"Total flagged cost by vendor — click to drill",
                  color="#dc3545",height=380,highlight=st.session_state.drill_vendor),
            use_container_width=True,on_select="rerun",key="vend_cost")
        cv2 = extract_click(ev_vco)
        if cv2 and cv2 in vcost:
            st.session_state.drill_vendor = cv2
            st.session_state.active_tab   = "vendors"

    # Vendor drill panel
    if st.session_state.drill_vendor and st.session_state.active_tab=="vendors":
        dv = st.session_state.drill_vendor
        drill_banner(f"Vendor: {dv}","drill_vendor")
        vdf = df[df["vendor"]==dv]
        vdf_i = (vdf.explode("issues_list").rename(columns={"issues_list":"issue"})
                 .query("issue.notna() and issue != ''"))
        vc1,vc2,vc3,vc4 = st.columns(4)
        vc1.metric("Bills flagged", len(vdf))
        vc2.metric("Unresolved",    (vdf["status"]=="Unresolved").sum())
        vc3.metric("Total cost",    f"${vdf['cost'].sum():,.0f}")
        vc4.metric("Unique issues", vdf_i["issue"].nunique())
        v2a,v2b = st.columns(2)
        with v2a:
            vi_ct = vdf_i["issue"].value_counts().to_dict()
            if vi_ct:
                ev_vi = st.plotly_chart(
                    bar_h(vi_ct,f"Issues — {dv}",color="#fd7e14",
                          height=max(200,len(vi_ct)*36+60)),
                    use_container_width=True,on_select="rerun",key="vd_issue")
                cv3 = extract_click(ev_vi)
                if cv3: st.session_state.drill_issue = cv3
        with v2b:
            vp_ct = (vdf[vdf["billing_period_label"].notna()]
                     .groupby("billing_period_label").size().sort_index().to_dict())
            if vp_ct:
                st.plotly_chart(bar_v(vp_ct,f"Bills by period — {dv}",
                                      color="#0d6efd",height=300),
                                use_container_width=True)
        render_bill_cards(vdf.sort_values("cost",ascending=False))

    # Summary table (clickable rows)
    st.divider()
    st.subheader("All vendors — summary")
    vsumm = (df.groupby("vendor")
               .agg(bills=("bill_id","count"),
                    unresolved=("status",lambda x:(x=="Unresolved").sum()),
                    total_cost=("cost","sum"),
                    unique_issues=("issues_list",lambda x:len(set(i for lst in x for i in lst))))
               .sort_values("bills",ascending=False).reset_index())
    vsumm.columns = ["Vendor","Bills","Unresolved","Total Cost ($)","Unique Issue Types"]
    vsumm["Total Cost ($)"] = vsumm["Total Cost ($)"].map("${:,.0f}".format)
    vs = st.dataframe(vsumm,use_container_width=True,hide_index=True,
                      on_select="rerun",selection_mode="single-row",key="vendor_table")
    vsr = vs.get("selection",{}).get("rows",[]) if vs else []
    if vsr:
        pv = vsumm.iloc[vsr[0]]["Vendor"]
        st.session_state.drill_vendor = pv
        st.session_state.active_tab   = "vendors"
        st.rerun()


# ══════════════════════════════════════════════════════════════════════════════
# TAB 4 — GHG PRIORITY QUEUE
# ══════════════════════════════════════════════════════════════════════════════
with t_ghg:
    st.subheader("🌿 GHG Emissions Priority Queue")

    if not has_usage:
        st.info(
            "Upload **Report-18** (Bill Line Item Report) alongside Report-27 to enable "
            "GHG usage analysis and priority scoring.\n\n"
            "Report-18 provides the usage quantity and unit of measure per bill, which is "
            "converted to a common energy unit for emissions impact scoring."
        )
    else:
        # ── Display unit selector ─────────────────────────────────────────────
        col_u, col_s, col_c = st.columns([2,2,4])
        with col_u:
            disp_unit = st.selectbox("Display unit", DISPLAY_UNITS, index=0, key="ghg_unit")
        with col_s:
            ghg_status = st.multiselect("Status filter",
                                         ["Resolved","Unresolved"],
                                         default=["Resolved","Unresolved"],
                                         key="ghq_status")
        with col_c:
            st.markdown("")
            st.caption(
                "Bills are scored by **GHG emissions impact**: "
                "unresolved flags with high energy usage and Scope 1/2 commodities rank highest. "
                "Usage is converted to kWh equivalent for cross-commodity comparison."
            )

        ghg_df = df[df["total_kwh_equivalent"].notna()].copy()
        if ghg_status:
            ghg_df = ghg_df[ghg_df["status"].isin(ghg_status)]

        if ghg_df.empty:
            st.warning("No GHG usage data matches the current filters.")
        else:
            # Priority score: unresolved × flag_risk × kwh_magnitude
            flag_risk = {"High":3,"Medium":2,"Low":1}
            ghg_df["flag_risk_score"] = ghg_df["primary_priority"].map(flag_risk).fillna(1)
            ghg_df["unresolved_mult"] = (ghg_df["status"]=="Unresolved").astype(int)*2 + 1
            ghg_df["priority_score"]  = (ghg_df["total_kwh_equivalent"]
                                          * ghg_df["flag_risk_score"]
                                          * ghg_df["unresolved_mult"])
            ghg_df["display_usage"]   = ghg_df["total_kwh_equivalent"].apply(
                                          lambda x: from_kwh(x, disp_unit))
            ghg_df = ghg_df.sort_values("priority_score", ascending=False)

            # ── Summary by GHG category ───────────────────────────────────────
            st.markdown("### Usage by GHG category")
            st.caption(f"Total displayed in **{disp_unit}**. Select a display unit above to change.")

            cat_summary = (ghg_df.groupby("ghg_category")
                           .agg(bills=("bill_id","count"),
                                total_usage=("display_usage","sum"),
                                unresolved=("status",lambda x:(x=="Unresolved").sum()))
                           .sort_values("total_usage",ascending=False)
                           .reset_index())
            cat_summary.columns = ["GHG Category","Bills","Total Usage","Unresolved Bills"]
            cat_summary["Total Usage"] = cat_summary["Total Usage"].apply(
                lambda x: f"{x:,.1f} {disp_unit}")

            ghg_scope_map = {v:GHG_SCOPE.get(v,"") for v in ghg_df["ghg_category"].unique()}
            cat_summary.insert(1,"Scope",cat_summary["GHG Category"].map(ghg_scope_map))
            st.dataframe(cat_summary,use_container_width=True,hide_index=True)

            # ── Usage bar chart ───────────────────────────────────────────────
            cat_bar = (ghg_df.groupby("ghg_category")["display_usage"]
                       .sum().sort_values(ascending=False).to_dict())
            if cat_bar:
                scope_colors = {"Electricity":"#0d6efd","Natural Gas":"#fd7e14",
                                "LPG / Propane":"#e67e22","Biomass":"#198754",
                                "District Heat / Steam":"#6f42c1",
                                "Diesel / Fuel Oil":"#dc3545","Gasoline":"#e74c3c",
                                "Aviation Fuel":"#c0392b","Coal":"#495057"}
                bar_colors = [scope_colors.get(k,"#888") for k in cat_bar]
                fig_cat = go.Figure(go.Bar(
                    x=list(cat_bar.values()), y=list(cat_bar.keys()),
                    orientation="h", marker_color=bar_colors,
                    text=[f"{v:,.0f}" for v in cat_bar.values()],
                    textposition="outside"))
                fig_cat.update_layout(
                    title=f"Flagged bill usage by commodity ({disp_unit})",
                    yaxis_autorange="reversed",
                    xaxis=dict(showgrid=True,gridcolor="#f0f0f0"),
                    height=280, margin=dict(l=10,r=80,t=40,b=10),
                    showlegend=False, **PT)
                st.plotly_chart(fig_cat,use_container_width=True)

            # ── Priority queue table ──────────────────────────────────────────
            st.markdown("### Priority queue — ranked by emissions impact")
            st.caption(
                "Score = Usage (kWh equiv.) × Flag risk (High=3, Med=2, Low=1) × "
                "Status multiplier (Unresolved=3×, Resolved=1×)"
            )

            pq_disp = ghg_df[[
                "bill_id","vendor","billing_period_label","status",
                "ghg_category","ghg_scope","display_usage","cost",
                "primary_priority","flag_issues","assigned_to","priority_score"
            ]].copy()
            pq_disp.columns = [
                "Bill ID","Vendor","Period","Status","Commodity","Scope",
                f"Usage ({disp_unit})","Cost ($)","Priority","Flag Issues",
                "Assigned To","Score"
            ]
            pq_disp[f"Usage ({disp_unit})"] = pq_disp[f"Usage ({disp_unit})"].apply(
                lambda x: f"{x:,.1f}")
            pq_disp["Cost ($)"]  = pq_disp["Cost ($)"].map("${:,.2f}".format)
            pq_disp["Score"]     = pq_disp["Score"].apply(lambda x: f"{x:,.0f}")

            pq_sel = st.dataframe(pq_disp,use_container_width=True,hide_index=True,
                                  height=480,on_select="rerun",
                                  selection_mode="single-row",key="pq_table")
            pq_rows = pq_sel.get("selection",{}).get("rows",[]) if pq_sel else []
            if pq_rows:
                st.markdown("---")
                bill_detail_panel(ghg_df.iloc[pq_rows[0]])

            buf2 = io.StringIO(); pq_disp.to_csv(buf2,index=False)
            st.download_button("⬇ Download priority queue as CSV",
                               buf2.getvalue(),"ghg_priority_queue.csv","text/csv")


# ══════════════════════════════════════════════════════════════════════════════
# TAB 5 — ACTION GUIDE
# ══════════════════════════════════════════════════════════════════════════════
with t_actions:
    st.subheader("Actionable Insights & Recommended Next Steps")
    st.caption("Sorted by priority (High → Medium → Low) then occurrence count. "
               "Each issue links to specific EnergyCAP navigation steps.")

    # Open flags
    open_df = df[df["status"]=="Unresolved"]
    if not open_df.empty:
        st.markdown("### 🔴 Open / Unresolved Flags")
        for _,row in open_df.iterrows():
            p   = row["primary_priority"]
            css = {"High":"action-high","Medium":"action-medium","Low":"action-low"}.get(p,"action-info")
            bc  = {"High":"red","Medium":"orange","Low":"green"}.get(p,"blue")
            badges = " ".join(badge(i,bc) for i in row["issues_list"])
            usage_str = ""
            if has_usage and pd.notna(row.get("total_kwh_equivalent")):
                usage_str = (f"&nbsp;·&nbsp; {row['total_kwh_equivalent']:,.0f} kWh"
                             f"{' · '+row['ghg_category'] if row.get('ghg_category') else ''}")
            st.markdown(f"""<div class="action-card {css}">
                <strong>Bill {row['bill_id']} — {row['vendor']}</strong>
                &nbsp;|&nbsp; ${row['cost']:,.2f}{usage_str}
                &nbsp;|&nbsp; Assigned: {row.get('assigned_to','—') or '—'}<br>{badges}
            </div>""", unsafe_allow_html=True)
        st.divider()

    # Per-issue guidance
    sorted_issues = sorted(
        [(i,c) for i,c in issue_counts.items() if i in FLAG_META],
        key=lambda x:(PRIORITY_ORDER.get(FLAG_META[x[0]]["priority"],1),-x[1]))
    seen_cats: set = set()
    for issue,cnt in sorted_issues:
        meta = FLAG_META[issue]
        cat  = meta["category"]
        if cat not in seen_cats:
            seen_cats.add(cat)
            st.markdown(f"### {cat}")
        p   = meta["priority"]
        css = {"High":"action-high","Medium":"action-medium","Low":"action-low"}.get(p,"action-info")
        bc  = {"High":"red","Medium":"orange","Low":"green"}.get(p,"blue")
        ib  = issues_exp[issues_exp["issue"]==issue]
        unr = (ib["status"]=="Unresolved").sum()
        topv= ib["vendor"].value_counts().idxmax() if not ib.empty else "—"
        unr_html = f'&nbsp;{badge("⚠ "+str(unr)+" unresolved","red")}' if unr else ""
        with st.container():
            st.markdown(f"""<div class="action-card {css}">
                <div style="display:flex;justify-content:space-between;align-items:flex-start;">
                    <div><strong>{issue}</strong>
                        &nbsp;{badge(p,bc)}&nbsp;{badge(str(cnt)+" occ"+("s" if cnt>1 else ""),"gray")}
                        {unr_html}</div>
                    <div style="font-size:12px;color:#6c757d">Top vendor: {topv}</div>
                </div>
                <p style="margin:8px 0 4px;font-size:13px;color:#555"><em>{meta['cause']}</em></p>
                <p style="margin:4px 0 2px;font-size:13px"><strong>Action:</strong> {meta['action']}</p>
                <p style="margin:2px 0 0;font-size:12px;color:#6c757d"><strong>In EnergyCAP:</strong> {meta['in_energycap']}</p>
            </div>""", unsafe_allow_html=True)
            matching = df[df["issues_list"].apply(lambda lst: issue in lst)]
            with st.expander(f"Show {len(matching)} bill{'s' if len(matching)!=1 else ''} with this issue"):
                render_bill_cards(matching.sort_values("cost",ascending=False))

    # Systemic insights
    st.divider()
    st.markdown("### 💡 Systemic Observations")
    insights = []
    rs = issue_counts.get("Rate schedule mismatch",0)
    sn = issue_counts.get("Serial number mismatch",0)
    if rs+sn > total_bills*0.4:
        insights.append(("action-info","High volume of import/configuration mismatches",
            f"{rs} rate schedule + {sn} serial number mismatches ({(rs+sn)/total_bills*100:.0f}% of occurrences). "
            "Consider a bulk meter configuration update in EnergyCAP rather than resolving bill-by-bill."))
    dup = (issue_counts.get("Duplicate bill",0)+issue_counts.get("Overlapping bill",0)
           +issue_counts.get("Multiple bills in period",0))
    if dup > 0:
        insights.append(("action-high",f"Potential duplicate payments — {dup} overlap/duplicate flags",
            "These carry the highest financial risk. Verify each before releasing to AP. "
            "Cross-reference with your payment system to confirm no duplicates were paid."))
    if vendor_counts:
        topvn = max(vendor_counts,key=vendor_counts.get)
        topvp = vendor_counts[topvn]/total_bills*100
        if topvp>20:
            insights.append(("action-medium",
                f"High concentration: {topvn} ({topvp:.0f}% of flagged bills)",
                f"{topvn} is disproportionately represented. Review import template and meter mappings."))
    if total_recovery==0 and total_bills>10:
        insights.append(("action-info","No cost recovery tracked ($0.00)",
            "Start logging Cost Recovery in EnergyCAP when billing errors are corrected. "
            "This quantifies the ROI of your flag review process."))
    for css,title,body in insights:
        st.markdown(f"""<div class="action-card {css}">
            <strong>{title}</strong>
            <p style="margin:6px 0 0;font-size:13px">{body}</p>
        </div>""", unsafe_allow_html=True)
    if not insights:
        st.success("No major systemic issues detected with current filters.")


# ══════════════════════════════════════════════════════════════════════════════
# TAB 6 — BILL DETAIL
# ══════════════════════════════════════════════════════════════════════════════
with t_detail:
    st.subheader("Bill Detail")
    c1,c2,c3,c4 = st.columns(4)
    with c1: search   = st.text_input("Search bill ID or account","")
    with c2: issue_f  = st.multiselect("Flag issue",sorted(issue_counts),key="det_issues")
    with c3: vendor_f = st.multiselect("Vendor",all_vendors,key="det_vendor")
    with c4: det_sort = st.selectbox("Sort by",SORT_OPTS,key="det_sort")

    ddf = df.copy()
    if search:
        ddf = ddf[ddf["bill_id"].str.contains(search,case=False)|
                  ddf["account"].str.contains(search,case=False,na=False)]
    if issue_f:
        ddf = ddf[ddf["issues_list"].apply(lambda lst: any(i in lst for i in issue_f))]
    if vendor_f: ddf = ddf[ddf["vendor"].isin(vendor_f)]
    dsc,dsa = SORT_MAP[det_sort]
    ddf = ddf.sort_values(dsc,ascending=dsa,na_position="last")

    disp2_cols = {"bill_id":"Bill ID","vendor":"Vendor","billing_period_label":"Period",
                  "cost":"Cost ($)","status":"Status","flag_issues":"Flag Issues",
                  "assigned_to":"Assigned To","days_to_resolve":"Days",
                  "ghg_category":"Commodity","total_kwh_equivalent":"kWh Equiv.",
                  "cost_recovery":"Recovery ($)"}
    show2 = [c for c in disp2_cols if c in ddf.columns]
    disp2 = ddf[show2].copy().rename(columns=disp2_cols)
    disp2["Cost ($)"]  = disp2["Cost ($)"].map("${:,.2f}".format)
    if "Recovery ($)" in disp2.columns:
        disp2["Recovery ($)"] = disp2["Recovery ($)"].map("${:,.2f}".format)
    if "kWh Equiv." in disp2.columns:
        disp2["kWh Equiv."] = disp2["kWh Equiv."].apply(
            lambda x: f"{x:,.0f}" if pd.notna(x) else "—")
    disp2["Days"] = disp2["Days"].apply(lambda x: f"{int(x)}d" if pd.notna(x) else "—")

    tbl2 = st.dataframe(disp2,use_container_width=True,hide_index=True,height=440,
                        on_select="rerun",selection_mode="single-row",key="detail_table")
    sel2 = tbl2.get("selection",{}).get("rows",[]) if tbl2 else []
    if sel2:
        st.markdown("---")
        bill_detail_panel(ddf.iloc[sel2[0]])

    buf3 = io.StringIO(); disp2.to_csv(buf3,index=False)
    st.download_button("⬇ Download as CSV",buf3.getvalue(),"bill_detail.csv","text/csv")
